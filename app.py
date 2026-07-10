import os
import time
import streamlit as st
from dotenv import load_dotenv

# Modules qui produisent du PPT ou de l'Excel (pas du Word)
_PPT_MODULES_BUY  = set()
_EXCEL_MODULES    = set()


def _generate_single_xlsx(mod_key: str, mod_label: str, text: str, company: str) -> bytes:
    """Génère un Excel simple (sans IA) depuis le markdown d'un module."""
    import io, re as _re2
    from openpyxl import Workbook as _WB
    from openpyxl.styles import Font as _Font, PatternFill as _Fill, Alignment as _Align, Border as _Bdr, Side as _Side

    wb = _WB()
    ws = wb.active
    ws.title = mod_label[:31]
    ws.sheet_view.showGridLines = False

    thin = _Side(style="thin", color="DDDDDD")
    bdr  = _Bdr(left=thin, right=thin, top=thin, bottom=thin)

    def _hcell(r, c, v):
        x = ws.cell(r, c, v)
        x.font  = _Font(name="Arial", bold=True, color="FFFFFF", size=9)
        x.fill  = _Fill("solid", fgColor="1F3864")
        x.alignment = _Align(horizontal="center", vertical="center", wrap_text=True)
        x.border = bdr
    def _dcell(r, c, v, even=True):
        x = ws.cell(r, c, v)
        x.font  = _Font(name="Arial", size=8, color="111111")
        x.fill  = _Fill("solid", fgColor="F2F4F8" if even else "FFFFFF")
        x.alignment = _Align(horizontal="left", vertical="center", wrap_text=True)
        x.border = bdr

    # En-tête classeur
    ws.merge_cells("A1:Z1")
    h = ws.cell(1, 1, f"{company} — {mod_label}")
    h.font = _Font(name="Arial", bold=True, size=12, color="FFFFFF")
    h.fill = _Fill("solid", fgColor="1F3864")
    h.alignment = _Align(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[1].height = 24

    # Parser le tableau Markdown
    lines = [l.strip() for l in text.split("\n")]
    table_lines = [l for l in lines if l.startswith("|")]
    
    if table_lines:
        row_idx = 3
        for i, line in enumerate(table_lines):
            if _re2.match(r"^[|][-| :]+[|]$", line):
                continue
            cells = [c.strip() for c in line.strip("|").split("|")]
            if i == 0:
                for ci, val in enumerate(cells, start=1):
                    _hcell(row_idx, ci, val)
                    ws.column_dimensions[chr(64+ci)].width = max(12, min(40, len(val)+4))
                ws.row_dimensions[row_idx].height = 22
            else:
                for ci, val in enumerate(cells, start=1):
                    _dcell(row_idx, ci, val, even=row_idx % 2 == 0)
                ws.row_dimensions[row_idx].height = 32
            row_idx += 1
    else:
        # Pas de tableau — écrire le texte brut ligne par ligne
        row_idx = 3
        for line in lines:
            if line:
                ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=6)
                x = ws.cell(row_idx, 1, line)
                x.font = _Font(name="Arial", size=8)
                x.alignment = _Align(wrap_text=True)
                ws.row_dimensions[row_idx].height = 18
                row_idx += 1

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _render_question_cards(step_key: str, step_info: dict, company: str):
    """
    Affiche les questions UNE PAR UNE (wizard) avec navigation Précédent / Suivant.
    Retourne (step_input, variables, ready) :
      - ready=False tant que toutes les questions n'ont pas été validées
      - ready=True quand l'utilisateur a cliqué "Valider" sur la dernière question
    """
    questions = step_info.get("questions", [])
    if not questions:
        return "", {}, True

    n       = len(questions)
    idx_key = f"q_idx_{step_key}"
    if idx_key not in st.session_state:
        st.session_state[idx_key] = 0

    current_idx = st.session_state[idx_key]

    # ── Toutes les questions validées ────────────────────────────────────────
    if current_idx >= n:
        variables  = {}
        step_input = ""
        for q in questions:
            qk    = q["key"]
            qtype = q["type"]
            if qtype == "chips_or_custom":
                chip = st.session_state.get(f"q_{step_key}_{qk}_pill")
                if q.get("multi"):
                    _sel = chip if isinstance(chip, list) else ([chip] if chip else [])
                    _real = [c for c in _sel if c != "Autre..."]
                    if "Autre..." in _sel:
                        _custom = st.session_state.get(f"q_{step_key}_{qk}_custom", "")
                        _real += [l.strip() for l in _custom.split("\n") if l.strip()]
                    val = " + ".join(_real) if _real else ""
                else:
                    val  = (st.session_state.get(f"q_{step_key}_{qk}_custom", "")
                            if chip == "Autre..." else chip or "")
            elif qtype == "text_or_file":
                mode = st.session_state.get(f"q_{step_key}_{qk}_mode", "✏️  Coller du texte")
                if mode == "📎  Uploader un fichier":
                    val = st.session_state.get(f"q_file_{step_key}_{qk}", "")
                else:
                    val = st.session_state.get(f"q_{step_key}_{qk}_text", "")
                step_input = val
            else:
                val = st.session_state.get(f"q_{step_key}_{qk}", "")
            variables[qk] = val

        if not step_input:
            parts = [
                f"Positionnement : {variables['positionnement']}" if variables.get("positionnement") else "",
                f"Catégories verticales :\n{variables['categories']}" if variables.get("categories") else "",
                f"Zone géographique : {variables['zone_geo']}" if variables.get("zone_geo") else "",
                f"Fourchette CA : {variables['fourchette_ca']}" if variables.get("fourchette_ca") else "",
                f"Entreprises à exclure :\n{variables['exclusions']}" if variables.get("exclusions") else "",
            ]
            step_input = "\n".join(p for p in parts if p)

        # Récap compact + lien pour modifier
        st.markdown(
            '<div style="display:flex;align-items:center;gap:10px;margin-bottom:4px;">'
            '<span style="font-size:0.85rem;color:#111111;font-weight:600;">✓ Questions renseignées</span>'
            '</div>',
            unsafe_allow_html=True,
        )
        if st.button("✏️ Modifier les réponses", key=f"q_reset_{step_key}"):
            st.session_state[idx_key] = 0
            st.rerun()

        return step_input, variables, True

    # ── Question courante ────────────────────────────────────────────────────
    q        = questions[current_idx]
    qk       = q["key"]
    label    = q["label"].replace("{company}", company)
    qtype    = q["type"]
    required = q.get("required", True)
    badge    = "" if required else '<span class="q-optional-badge">optionnel</span>'

    # Indicateur de progression (points)
    dots = ""
    for i in range(n):
        if i < current_idx:
            dots += '<div style="width:8px;height:8px;border-radius:50%;background:#111111;flex-shrink:0;"></div>'
        elif i == current_idx:
            dots += '<div style="width:10px;height:10px;border-radius:50%;background:#111111;flex-shrink:0;"></div>'
        else:
            dots += '<div style="width:8px;height:8px;border-radius:50%;background:#D5D5D5;flex-shrink:0;"></div>'
    st.markdown(
        f'<div style="display:flex;align-items:center;gap:7px;margin-bottom:6px;">'
        f'{dots}'
        f'<span style="font-size:0.76rem;color:#9CA3AF;margin-left:3px;">{current_idx + 1} / {n}</span>'
        f'</div>',
        unsafe_allow_html=True,
    )

    with st.container(border=True):
        st.markdown(f'<p class="q-card-label">{label}{badge}</p>', unsafe_allow_html=True)

        # ── Texte court ──────────────────────────────────────────────────────
        if qtype == "text_input":
            st.text_input(
                label, placeholder=q.get("hint", ""),
                key=f"q_{step_key}_{qk}", label_visibility="collapsed",
            )

        # ── Texte long ───────────────────────────────────────────────────────
        elif qtype == "textarea":
            st.text_area(
                label, placeholder=q.get("hint", ""), height=120,
                key=f"q_{step_key}_{qk}", label_visibility="collapsed",
            )

        # ── Chips + "Autre" ──────────────────────────────────────────────────
        elif qtype == "chips_or_custom":
            _q_multi = q.get("multi", False)
            options = q.get("options", []) + ["Autre..."]
            chip = st.pills(
                label, options=options,
                selection_mode="multi" if _q_multi else "single",
                key=f"q_{step_key}_{qk}_pill", label_visibility="collapsed",
            )
            _has_autre = ("Autre..." in (chip or [])) if _q_multi else (chip == "Autre...")
            if _has_autre:
                st.text_input(
                    "Précisez", placeholder=q.get("hint", ""),
                    key=f"q_{step_key}_{qk}_custom",
                )

        # ── Texte ou fichier ─────────────────────────────────────────────────
        elif qtype == "text_or_file":
            mode = st.pills(
                "Mode",
                options=["✏️  Coller du texte", "📎  Uploader un fichier"],
                selection_mode="single", default="✏️  Coller du texte",
                key=f"q_{step_key}_{qk}_mode", label_visibility="collapsed",
            )
            if mode != "📎  Uploader un fichier":
                st.text_area(
                    label, placeholder=q.get("hint_text", ""), height=180,
                    key=f"q_{step_key}_{qk}_text", label_visibility="collapsed",
                )
            else:
                up = st.file_uploader(
                    "Document", type=["pdf", "docx", "txt", "md"],
                    accept_multiple_files=True,
                    key=f"q_{step_key}_{qk}_file",
                    label_visibility="collapsed",
                    help="Formats : .docx, .pdf, .txt — max 3 fichiers",
                )
                if up:
                    from document_extractor import extract_text
                    content = ""
                    for uf in up[:3]:
                        content += f"\n\n--- {uf.name} ---\n{extract_text(uf)}"
                    st.session_state[f"q_file_{step_key}_{qk}"] = content

        # ── Navigation ───────────────────────────────────────────────────────
        st.markdown("<div style='height:4px'></div>", unsafe_allow_html=True)
        col_prev, col_next = st.columns([1, 2])

        with col_prev:
            if current_idx > 0:
                if st.button("← Retour", key=f"q_prev_{step_key}_{current_idx}",
                             use_container_width=True):
                    st.session_state[idx_key] = current_idx - 1
                    st.rerun()

        with col_next:
            btn_label = "Valider →" if current_idx == n - 1 else "Suivant →"
            if st.button(btn_label, key=f"q_next_{step_key}_{current_idx}",
                         type="primary", use_container_width=True):
                st.session_state[idx_key] = current_idx + 1
                st.rerun()

    return "", {}, False


def _export_button(result: str, company: str, step_key: str, step_info: dict):
    """Affiche le bon bouton de téléchargement selon le type de module."""
    fname_base = f"{company.replace(' ', '_').lower()}_{step_info['num']}_{step_info['title'].replace(' ', '_')}"
    try:
        if step_key in _PPT_MODULES_BUY:
            from export_pptx import generate_buy_pptx, try_exec_pptx_code
            data = try_exec_pptx_code(result) or generate_buy_pptx(result, company, step_info.get("title", "Cibles"))
            st.download_button(
                "📊 Télécharger en PowerPoint",
                data=data,
                file_name=f"{fname_base}.pptx",
                mime="application/vnd.openxmlformats-officedocument.presentationml.presentation",
            )
        elif step_key in _EXCEL_MODULES:
            from export_pptx import generate_screening_xlsx
            data = generate_screening_xlsx(result, company)
            st.download_button(
                "📥 Télécharger en Excel",
                data=data,
                file_name=f"{fname_base}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        else:
            from word_generator import generate_word
            st.download_button(
                "📝 Télécharger en Word",
                data=generate_word(result, company, "fiche"),
                file_name=f"{fname_base}.docx",
                mime="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
            )
    except Exception as _export_err:
        st.warning(f"⚠️ Export indisponible : {_export_err}")

load_dotenv(os.path.join(os.path.dirname(__file__), ".env"), override=True)

st.set_page_config(
    page_title="Screening M&A",
    page_icon="🔍",
    layout="wide",
)

# ── API key helpers ────────────────────────────────────────────────────────────

def _research_company(company: str, anthropic_key: str, tavily_key: str) -> dict:
    """
    Recherche agentique (Sonnet + Tavily tool-use loop) — même moteur que les analyses rapides.
    Retourne un dict avec description, faits, positionnement_options, categories_options.
    """
    import anthropic as _ant, json, re as _re, os as _os, time as _time
    from tools import TOOL_DEFINITIONS, execute_tool

    _os.environ["ANTHROPIC_API_KEY"] = anthropic_key
    _os.environ["TAVILY_API_KEY"]    = tavily_key

    client = _ant.Anthropic(api_key=anthropic_key)

    system = f"""Tu es un analyste M&A senior. Fais 3 à 5 recherches web ciblées sur l'entreprise "{company}" pour collecter :
- Secteur précis, positionnement marché
- Chiffre d'affaires et effectifs (les plus récents)
- Implantations géographiques
- Spécialités, offres, clients types
- Date de création, actionnariat si disponible

Une fois les recherches faites, réponds UNIQUEMENT avec un JSON valide (aucun autre texte) structuré ainsi :
{{
  "description": "1 phrase précise : secteur, positionnement, localisation, taille",
  "faits": ["fait 1", "fait 2", "fait 3", "fait 4", "fait 5"],
  "positionnement_options": ["formulation 1 (max 9 mots)", "formulation 2", "formulation 3", "formulation 4"],
  "categories_options": ["V1 — Libellé", "V2 — Libellé", "V3 — Libellé", "V4 — Libellé", "V5 — Libellé"]
}}"""

    messages = [{"role": "user", "content": f"Recherche complète sur l'entreprise : {company}"}]
    full_text = ""

    for _ in range(12):  # max 12 tours (≈ 5-6 appels web + rédaction)
        for attempt in range(3):
            try:
                response = client.messages.create(
                    model="claude-sonnet-4-6",
                    max_tokens=2000,
                    system=system,
                    messages=messages,
                    tools=TOOL_DEFINITIONS,
                )
                break
            except Exception:
                if attempt == 2:
                    raise
                _time.sleep(5)

        tool_uses = []
        for block in response.content:
            if hasattr(block, "text"):
                full_text = block.text
            elif getattr(block, "type", None) == "tool_use":
                tool_uses.append(block)

        if response.stop_reason == "end_turn" or not tool_uses:
            break

        tool_results = [
            {"type": "tool_result", "tool_use_id": tu.id, "content": execute_tool(tu.name, tu.input)}
            for tu in tool_uses
        ]
        messages.append({"role": "assistant", "content": response.content})
        messages.append({"role": "user", "content": tool_results})

    m = _re.search(r"\{[^{}]*\{.*?\}[^{}]*\}", full_text, _re.DOTALL)
    if not m:
        m = _re.search(r"\{.*\}", full_text, _re.DOTALL)
    if m:
        try:
            return json.loads(m.group())
        except Exception:
            pass
    return {}


def get_key(name):
    """Read API keys from Streamlit Cloud (st.secrets) or local .env file."""
    try:
        return st.secrets[name]
    except Exception:
        return os.environ.get(name, "")

anthropic_key = get_key("ANTHROPIC_API_KEY")
tavily_key    = get_key("TAVILY_API_KEY")

# ── Styling ────────────────────────────────────────────────────────────────────
st.markdown("""
<style>
/* ── Fonts ── */
@import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700;800&display=swap');

html, body, [class*="css"] {
    font-family: 'Inter', 'Helvetica Neue', Arial, sans-serif;
    color: #111111;
}

/* ── Animations ── */
@keyframes fadeIn { from { opacity:0; transform:translateY(8px); } to { opacity:1; transform:translateY(0); } }
@keyframes pulse  { 0%,100%{opacity:1;transform:scale(1)} 50%{opacity:.3;transform:scale(1.4)} }
@keyframes progressSlide { from{width:0%} to{width:100%} }

/* ── Fond général ── */

.main .block-container {
    max-width: 900px;
    padding-top: 2rem;
    margin: 0 auto;
    background: transparent;
    animation: fadeIn 0.3s ease both;
}

/* ── Sidebar ── */
section[data-testid="stSidebar"] {
    background: #FFFFFF;
    border-right: 1px solid #E5E5E5;
    box-shadow: none;
}
section[data-testid="stSidebar"] * { color: #111111 !important; }
section[data-testid="stSidebar"] hr { border-color: #E5E5E5 !important; }
section[data-testid="stSidebar"] .sidebar-logo {
    font-size: 1.2rem; font-weight: 700; letter-spacing: -0.3px;
    color: #111111 !important;
}
section[data-testid="stSidebar"] .sidebar-sources {
    font-size: 0.82rem; line-height: 1.9; color: #6B7280 !important;
}
section[data-testid="stSidebar"] button {
    background: #FFFFFF !important;
    border: 1px solid #E5E5E5 !important;
    color: #111111 !important;
}
section[data-testid="stSidebar"] button:hover { background: #F0F0F0 !important; }

/* ── Markdown ── */
.stMarkdown p, .stMarkdown li, .stMarkdown td, .stMarkdown th {
    color: #111111 !important; font-size: 0.94rem; line-height: 1.75;
}
.stMarkdown h1 { color:#111111!important; font-size:1.4rem!important; font-weight:700!important; border-bottom:1px solid #E5E5E5; padding-bottom:8px; margin-top:24px!important; }
.stMarkdown h2 { color:#111111!important; font-size:1.15rem!important; font-weight:700!important; margin-top:20px!important; }
.stMarkdown h3 { color:#333333!important; font-size:1.0rem!important; font-weight:600!important; }
.stMarkdown strong { color:#111111!important; font-weight:600!important; }
.stMarkdown em { color:#555555!important; }
.stMarkdown code { background:#F5F5F5!important; color:#111111!important; border-radius:4px; padding:1px 5px; }
.stMarkdown blockquote { border-left:2px solid #111111!important; padding-left:12px; color:#555555!important; }
.stMarkdown hr { border-color:#E5E5E5!important; margin:16px 0!important; }

/* ── Tableaux ── */
.stMarkdown table { width:100%; border-collapse:collapse; margin:12px 0; }
.stMarkdown th { background:#111111!important; color:#FFFFFF!important; font-weight:600; padding:10px 14px; text-align:left; }
.stMarkdown td { color:#111111!important; padding:8px 14px; border-bottom:1px solid #E5E5E5; }
.stMarkdown tr:nth-child(even) td { background:#F9F9F9!important; }
.stMarkdown tr:hover td { background:#F3F3F3!important; }

/* ── Titre principal ── */
.main-title { font-size:2rem; font-weight:700; color:#111111; letter-spacing:-0.5px; margin-bottom:0.2rem; }
.main-title span { color:#111111; }
.main-subtitle { font-size:0.95rem; color:#6B7280; margin-bottom:2rem; line-height:1.6; }

/* ── Cartes livrables ── */
.card-btn {
    background: #FFFFFF; border: 1px solid #E5E5E5; border-radius: 12px;
    padding: 18px 20px; cursor: pointer; transition: all 0.15s ease;
    width: 100%; text-align: left;
}
.card-btn:hover { border-color:#111111; box-shadow:0 4px 14px rgba(0,0,0,0.08); transform:translateY(-2px); }
.card-btn.selected { background:#F7F7F7; border:1.5px solid #111111; }
.card-icon  { font-size:1.4rem; display:block; margin-bottom:8px; }
.card-title { font-size:0.95rem; font-weight:600; color:#111111; display:block; margin-bottom:4px; }
.card-desc  { font-size:0.79rem; color:#6B7280; line-height:1.5; }

/* ── Barre de progression ── */
.progress-topbar { height:2px; background:#E5E5E5; border-radius:1px; margin-bottom:20px; overflow:hidden; }
.progress-topbar-fill { height:100%; background:#111111; border-radius:1px; animation:progressSlide 40s linear forwards; }

.progress-container {
    background:#FFFFFF; border:1px solid #E5E5E5; border-radius:12px;
    padding:24px 28px; margin:16px 0;
}
.progress-header  { display:flex; align-items:center; gap:12px; margin-bottom:20px; }
.progress-title   { font-size:1.0rem; font-weight:700; color:#111111; }
.progress-company { font-size:0.87rem; color:#6B7280; }
.progress-deliverable {
    display:inline-block; background:#F5F5F5; border:1px solid #E5E5E5;
    color:#333333; font-weight:600; font-size:0.8rem;
    padding:3px 10px; border-radius:20px; margin-top:2px;
}
.step-row           { display:flex; align-items:center; gap:10px; padding:8px 0; font-size:0.9rem; color:#6B7280; border-bottom:1px solid #F0F0F0; }
.step-row:last-of-type { border-bottom:none; }
.step-row.done      { color:#111111; }
.step-row.active    { color:#111111; font-weight:600; }
.step-row.pending   { color:#C0C0C0; }
.step-icon          { font-size:1.0rem; min-width:24px; }
.pulse-dot { display:inline-block; width:7px; height:7px; background:#111111; border-radius:50%; margin-left:6px; animation:pulse 1.4s ease-in-out infinite; vertical-align:middle; }
.searching-label { font-size:0.82rem; color:#B0B0B0; margin-top:16px; display:flex; align-items:center; gap:6px; }

/* ── Résultat ── */
.result-box { background:#FFFFFF; border:1px solid #E5E5E5; border-radius:12px; padding:28px 32px; margin-top:16px; }
.result-header { font-size:1.0rem; font-weight:700; color:#111111; margin-bottom:14px; }
.company-badge {
    display:inline-flex; align-items:center; gap:6px;
    background:#F5F5F5; border:1px solid #E5E5E5;
    color:#111111; font-weight:700; font-size:1.0rem;
    padding:6px 14px; border-radius:8px; margin-bottom:10px;
}

/* ── Boutons Streamlit ── */
div[data-testid="stButton"] > button {
    border-radius:8px; font-weight:500; font-size:0.88rem;
    transition:all 0.15s ease;
    background:#FFFFFF; border:1px solid #E5E5E5; color:#111111;
}
div[data-testid="stButton"] > button:hover { border-color:#111111; }
div[data-testid="stButton"] > button[kind="primary"] { background:#111111; border:none; color:#FFFFFF; }
div[data-testid="stButton"] > button[kind="primary"]:hover { background:#333333; }
div[data-testid="stDownloadButton"] > button {
    border-radius:8px; font-size:0.88rem; font-weight:500;
    background:#FFFFFF; border:1px solid #E5E5E5; color:#111111;
    transition:all 0.15s ease;
}
div[data-testid="stDownloadButton"] > button:hover { border-color:#111111; }

/* ── Expander ── */
details { border:1px solid #E5E5E5!important; border-radius:10px!important; margin-bottom:12px!important; background:#FFFFFF!important; }
details summary { font-weight:600!important; color:#111111!important; }

/* ── Inputs ── */
div[data-testid="stTextInput"] input, div[data-testid="stTextArea"] textarea {
    background:#FFFFFF!important; border:1px solid #E5E5E5!important;
    border-radius:8px!important; color:#111111!important;
}
div[data-testid="stTextInput"] input:focus, div[data-testid="stTextArea"] textarea:focus {
    border-color:#111111!important; box-shadow:0 0 0 2px rgba(0,0,0,0.08)!important;
}

/* ── File uploader ── */
div[data-testid="stFileUploader"] {
    background:#FAFAFA!important; border:1.5px dashed #E5E5E5!important; border-radius:10px!important;
}

/* ── Tabs ── */
div[data-testid="stTabs"] button { color:#6B7280!important; font-weight:500!important; }
div[data-testid="stTabs"] button[aria-selected="true"] { color:#111111!important; font-weight:700!important; border-bottom:2px solid #111111!important; }
button[data-baseweb="tab"] { font-size:0.88rem!important; font-weight:600!important; }

/* ── Badges fichiers ── */
.file-badge {
    display:inline-flex; align-items:center; gap:6px;
    background:#F5F5F5; border:1px solid #E5E5E5; color:#333333;
    font-size:0.8rem; font-weight:500; padding:4px 10px; border-radius:20px; margin:4px 4px 4px 0;
}
.file-badge-name { max-width:160px; overflow:hidden; text-overflow:ellipsis; white-space:nowrap; }

/* ── Labels section ── */
.section-label {
    font-size:0.75rem; font-weight:600; color:#9CA3AF;
    text-transform:uppercase; letter-spacing:0.07em; margin-bottom:8px;
}

/* ── Spinner ── */
div[data-testid="stSpinner"] > div { font-size:0.9rem; color:#6B7280; }

/* ── Header Streamlit ── */
header[data-testid="stHeader"] { box-shadow:none!important; border-bottom:none!important; }
footer { visibility:hidden!important; }

/* ── Scrollbar ── */
::-webkit-scrollbar { width:5px; }
::-webkit-scrollbar-track { background:#FAFAFA; }
::-webkit-scrollbar-thumb { background:#D5D5D5; border-radius:3px; }
::-webkit-scrollbar-thumb:hover { background:#A0A0A0; }

/* ── Segmented control email ── */
.etype-btn-selected > div > button { background:#111111!important; border-color:#111111!important; color:#FFFFFF!important; }
.etype-btn-idle > div > button { background:#FFFFFF!important; border:1px solid #E5E5E5!important; color:#333333!important; }
.etype-btn-idle > div > button:hover { border-color:#111111!important; color:#111111!important; }

/* ── Question cards ── */

.q-optional-badge { font-size:0.72rem; font-weight:400; color:#9CA3AF; margin-left:8px; text-transform:uppercase; letter-spacing:0.05em; }

/* ── Fond général ── */
[data-testid="stAppViewContainer"] { background: #D0D0D0; }

/* ── Header ── */
header[data-testid="stHeader"] {
    background:#D0D0D0!important; box-shadow:none!important; border-bottom:none!important;
}

/* ── Cacher la barre décorative Streamlit ── */
[data-testid="stDecoration"] { display:none!important; }

/* ── st.pills → chips (fond clair, texte sombre) ── */
div[data-testid="stPills"] > label { display:none!important; }
button[data-testid="stPillsOptionButton"] {
    border:1px solid #D0D0D0!important; border-radius:100px!important;
    background:#F5F5F5!important; color:#333333!important;
    font-size:0.87rem!important; padding:7px 18px!important;
    margin:0 6px 6px 0!important; font-weight:400!important; box-shadow:none!important;
    transition:all 0.12s!important;
}
button[data-testid="stPillsOptionButton"]:hover {
    border-color:#555555!important;
}
button[data-testid="stPillsOptionButton"][aria-pressed="true"] {
    background:#111111!important; border-color:#111111!important; color:#FFFFFF!important;
}

/* ── Question card label ── */
.q-card-label { font-size:0.94rem; font-weight:500; color:#111111; margin:4px 0 12px 0; line-height:1.5; }
</style>
<script>
(function() {
  function disableAutocorrect() {
    document.querySelectorAll('input[type="text"], textarea').forEach(function(el) {
      el.setAttribute('autocorrect', 'off');
      el.setAttribute('autocomplete', 'off');
      el.setAttribute('autocapitalize', 'off');
      el.setAttribute('spellcheck', 'false');
    });
  }
  disableAutocorrect();
  new MutationObserver(disableAutocorrect).observe(document.body, { childList: true, subtree: true });
})();
</script>
""", unsafe_allow_html=True)

# ── Deliverable type definitions ──────────────────────────────────────────────
DELIVERABLES = [
    {
        "key":   "fiche",
        "icon":  "🏢",
        "title": "Fiche Entreprise",
        "desc":  "Identité, activité, CA, effectifs, clients, actionnariat",
        "steps": ["Analyse de l'entreprise"],
    },
    {
        "key":   "benchmark",
        "icon":  "⚔️",
        "title": "Benchmark Concurrents",
        "desc":  "Cartographie des acteurs avec fiches standardisées",
        "steps": ["Analyse de l'entreprise", "Cartographie des concurrents"],
    },
    {
        "key":   "manda",
        "icon":  "📰",
        "title": "Note M&A & Secteur",
        "desc":  "Transactions récentes, valorisations, fédérations, salons",
        "steps": ["Actualités M&A", "Qualification du secteur"],
    },
    {
        "key":   "geo",
        "icon":  "🌍",
        "title": "Analyse Géographique",
        "desc":  "Opportunités d'expansion internationale et acteurs locaux",
        "steps": ["Analyse de l'entreprise", "Analyse géographique"],
    },
]

DELIVERABLE_BY_KEY = {d["key"]: d for d in DELIVERABLES}

# Estimated duration in seconds per deliverable type (used for progress timer)
ESTIMATED_SECONDS = {
    "fiche":     90,    # ~1.5 min
    "benchmark": 210,   # ~3.5 min
    "manda":     180,   # ~3 min
    "geo":       180,   # ~3 min
}

# ── Screening Buy Side — définitions ─────────────────────────────────────────────────

BUY_SIDE_STEPS = [
    {
        "key":         "buy_01_carto_verticale",
        "num":         "1a",
        "title":       "Chaîne de valeur",
        "desc":        "Cartographie verticale — acteurs amont / aval et position de l'acquéreur",
        "web_search":  True,
        "needs_input": True,
        "input_label": "Présentation de la société",
        "input_hint":  "",
        "questions": [
            {
                "key":       "input_societe",
                "label":     "Informations sur {company}",
                "type":      "text_or_file",
                "hint_text": "Activité cœur, offres / technologies, segments clients, canaux de vente, zones géographiques, modèle économique (marque propre, OEM, MDD…), effectifs, CA approximatif.",
                "required":  True,
            },
        ],
    },
    {
        "key":         "buy_02_carto_horizontale",
        "num":         "1b",
        "title":       "Segments concurrentiels",
        "desc":        "Cartographie horizontale — 2 à 5 segments de marché et rivalité",
        "web_search":  True,
        "needs_input": True,
        "input_label": "Présentation de la société",
        "input_hint":  "",
        "questions": [
            {
                "key":       "input_societe",
                "label":     "Informations sur {company}",
                "type":      "text_or_file",
                "hint_text": "Activité cœur, offres / technologies, segments clients, canaux de vente, zones géographiques, modèle économique (marque propre, OEM, MDD…), effectifs, CA approximatif.",
                "required":  True,
            },
        ],
    },
    {
        "key":         "buy_03_recherche_cibles",
        "num":         "2",
        "title":       "Long-list de cibles",
        "desc":        "30 cibles d'acquisition potentielles avec fit stratégique et sources",
        "web_search":  True,
        "needs_input": True,
        "input_label": "Contexte de l'acquéreur",
        "input_hint":  "",
        "questions": [
            {
                "key":      "positionnement",
                "label":    "Quel est le positionnement de {company} ?",
                "type":     "text_input",
                "hint":     "Ex : intégrateur industriel en électricité, automatisme, mécanisation et maintenance",
                "required": True,
            },
            {
                "key":      "categories",
                "label":    "Quelles catégories verticales cibler ?",
                "type":     "textarea",
                "hint":     "Copiez les catégories issues de la cartographie verticale (étape 1a).\nEx : V4 — Intégrateurs de systèmes et lignes de production",
                "required": True,
            },
            {
                "key":      "zone_geo",
                "label":    "Quelle zone géographique pour les cibles ?",
                "type":     "chips_or_custom",
                "options":  ["France", "France + Belgique", "France + Europe", "Europe"],
                "hint":     "Autre zone...",
                "required": True,
            },
            {
                "key":      "fourchette_ca",
                "label":    "Quelle fourchette de CA pour les cibles ?",
                "type":     "chips_or_custom",
                "options":  ["2–5 M€", "5–20 M€", "10–50 M€", "20–100 M€"],
                "hint":     "Ex : 3–15 M€",
                "required": True,
            },
            {
                "key":      "exclusions",
                "label":    "Entreprises à exclure ?",
                "type":     "textarea",
                "hint":     "Laissez vide si aucune exclusion. Une société par ligne.",
                "required": False,
            },
        ],
    },
]

BUY_SIDE_STEPS_BY_KEY = {s["key"]: s for s in BUY_SIDE_STEPS}

# Questions unifiées pour le wizard M&A (tous modules confondus)
MA_BUY_WIZARD = [
    {
        "key":      "positionnement",
        "label":    "Quel est le positionnement de {company} en une phrase ?",
        "type":     "text_input",
        "hint":     "Ex : intégrateur industriel en électricité, automatisme et maintenance",
        "required": True,
    },
    {
        "key":      "categories",
        "label":    "Quelles catégories verticales de cibles visez-vous ?",
        "type":     "textarea",
        "hint":     "Ex : V4 — Intégrateurs de systèmes et lignes de production\nV7 — Maintenance industrielle",
        "required": True,
    },
    {
        "key":      "zone_geo",
        "label":    "Zone géographique cible pour les acquisitions",
        "type":     "chips_or_custom",
        "options":  ["France uniquement", "France + Belgique", "France + DACH", "Europe"],
        "hint":     "Autre zone...",
        "required": True,
    },
    {
        "key":      "fourchette_ca",
        "label":    "Fourchette de chiffre d'affaires des cibles",
        "type":     "chips_or_custom",
        "options":  ["< 5 M€", "5–20 M€", "20–50 M€", "50–100 M€"],
        "hint":     "Autre fourchette...",
        "required": True,
    },
    {
        "key":      "exclusions",
        "label":    "Entreprises à exclure explicitement",
        "type":     "textarea",
        "hint":     "Laissez vide si aucune. Une société par ligne.",
        "required": False,
    },
]

# ── Session state initialization ──────────────────────────────────────────────
def init_state():
    defaults = {
        "screen":           1,
        "company":          "",
        "deliverable_type": "fiche",
        "context":          "",
        "result_text":      "",
        "current_step":     "",
        "steps_done":       [],
        "email_target":     "",
        "generated_email":  "",
        "found_executives": [],
        "email_type":       "rachat",
        # Screening Buy Side
        "ma_universe":      "",   # "buy" or "sell"
        "ma_step_key":      "",   # current module key
        "ma_company":       "",
        "ma_sector":        "",   # secteur d'activité — contexte clé pour les recherches
        "ma_step_result":   {},   # dict: step_key -> result text
        "ma_running":       False,
    }
    for k, v in defaults.items():
        if k not in st.session_state:
            st.session_state[k] = v

init_state()

# ── Sidebar ───────────────────────────────────────────────────────────────────
with st.sidebar:
    st.markdown(
        '<div class="sidebar-logo">🔍 M&A <span class="sidebar-accent">Screening</span></div>',
        unsafe_allow_html=True,
    )
    st.markdown("---")
    st.markdown('<div class="section-label">Sources utilisées</div>', unsafe_allow_html=True)
    st.markdown(
        '<div class="sidebar-sources">'
        "• Pappers, Societe.com, Infogreffe<br>"
        "• CFnews, Les Echos, BFM Business<br>"
        "• LinkedIn, Crunchbase, Dealroom<br>"
        "• INSEE, data.gouv.fr<br>"
        "• Kompass, Corporama"
        "</div>",
        unsafe_allow_html=True,
    )
    st.markdown("---")
    if st.session_state.screen != 1:
        if st.button("← Nouvelle analyse", use_container_width=True):
            for k in ["screen", "company", "deliverable_type", "context", "result_text",
                      "current_step", "steps_done", "ma_universe", "ma_company",
                      "ma_step_key", "ma_running"]:
                if k == "screen":
                    st.session_state[k] = 1
                elif k in ("steps_done",):
                    st.session_state[k] = []
                elif k == "ma_step_result":
                    st.session_state[k] = {}
                else:
                    st.session_state[k] = ""
            st.session_state.ma_step_result = {}
            st.rerun()

    st.markdown("---")
    st.markdown('<div class="section-label">Inviter un consultant</div>', unsafe_allow_html=True)
    invite_email = st.text_input(
        "Email du consultant",
        placeholder="prenom.nom@cabinet.fr",
        label_visibility="collapsed",
        key="invite_email_field",
    )
    if invite_email and "@" in invite_email:
        subject = "Accès à l'outil M&A Screening IA"
        body = (
            f"Bonjour,\n\n"
            f"Je t'invite à utiliser M&A Screening IA, un outil d'intelligence économique "
            f"qui permet d'analyser une entreprise en profondeur en quelques minutes "
            f"(cartographie concurrentielle, actualités M&A, qualification sectorielle).\n\n"
            f"Accède à l'outil ici : https://screening-ma.streamlit.app\n\n"
            f"Bonne analyse,"
        )
        import urllib.parse
        mailto = f"mailto:{invite_email}?subject={urllib.parse.quote(subject)}&body={urllib.parse.quote(body)}"
        st.markdown(
            f'<a href="{mailto}" target="_blank" style="'
            'display:block; text-align:center; background:#F5F5F5; '
            'border:1px solid #CCCCCC; border-radius:8px; '
            'color:#333333 !important; font-size:0.85rem; font-weight:600; '
            'padding:8px 0; margin-top:6px; text-decoration:none; cursor:pointer;">'
            'Envoyer l\'invitation</a>',
            unsafe_allow_html=True,
        )
    elif invite_email:
        st.markdown(
            '<div style="color:#E88; font-size:0.78rem; margin-top:4px;">Email invalide</div>',
            unsafe_allow_html=True,
        )

    st.markdown("---")
    st.markdown(
        '<div class="section-label">À propos</div>',
        unsafe_allow_html=True,
    )
    st.markdown(
        '<div class="sidebar-sources">'
        "<strong style='color:#CCCCCC;'>M&A Screening IA</strong><br>"
        "Outil d'intelligence économique alimenté par l'IA pour les professionnels du M&A.<br><br>"
        "Analyse d'entreprises, cartographie concurrentielle, actualités sectorielle et prise de contact — en quelques minutes.<br><br>"
        "<a href='mailto:contact@screening-ma.fr' style='color:#9CA3AF;'>contact@screening-ma.fr</a>"
        "</div>",
        unsafe_allow_html=True,
    )
    st.markdown(
        '<div class="sidebar-sources" style="margin-top:10px;color:#9CA3AF;font-size:0.75rem;">'
        "v1.0 · 2026 · Tous droits réservés"
        "</div>",
        unsafe_allow_html=True,
    )

# ══════════════════════════════════════════════════════════════════════════════
# SCREEN 1 — INPUT
# ══════════════════════════════════════════════════════════════════════════════
if st.session_state.screen == 1:

    st.markdown(
        '<div class="main-title">Screening <span>M&A</span></div>',
        unsafe_allow_html=True,
    )
    st.markdown(
        '<div class="main-subtitle">Analysez une entreprise en profondeur grâce à l\'intelligence artificielle et aux données publiques.</div>',
        unsafe_allow_html=True,
    )

    tab_standard, tab_mission = st.tabs(["🔍 Analyses rapides", "💼 Screening Buy Side"])

    # ── TAB 2 : Screening Buy Side ────────────────────────────────────────────────
    with tab_mission:
        st.markdown(
            '<p style="font-size:0.82rem;color:#9CA3AF;margin-bottom:6px;font-weight:600;'
            'text-transform:uppercase;letter-spacing:0.06em;">Nom de l\'acquéreur</p>',
            unsafe_allow_html=True,
        )
        with st.form("ma_start_form", border=False):
            buy_company = st.text_input(
                "Acquéreur",
                placeholder="Ex : Milliris, Acuitis...",
                label_visibility="collapsed",
                key="buy_company_input",
            )
            st.markdown(
                '<p style="font-size:0.82rem;color:#9CA3AF;margin-top:10px;margin-bottom:4px;'
                'font-weight:600;text-transform:uppercase;letter-spacing:0.06em;">'
                'Documents de contexte <span style="font-weight:400;text-transform:none;'
                'letter-spacing:0;font-size:0.75rem;">— facultatif</span></p>',
                unsafe_allow_html=True,
            )
            ma_docs_upload = st.file_uploader(
                "Documents",
                type=["pdf", "docx", "txt", "md", "xlsx", "csv"],
                accept_multiple_files=True,
                label_visibility="collapsed",
                key="ma_start_docs",
                help="Plaquette, rapport annuel, mémo… L'IA les utilisera pendant la mission.",
            )
            submitted = st.form_submit_button(
                "Commencer →", type="primary", use_container_width=True
            )
        if submitted:
            if not buy_company.strip():
                st.warning("⚠️ Entrez le nom de l'acquéreur.")
            elif not anthropic_key or not tavily_key:
                st.error("🔑 Clés API manquantes.")
            else:
                # Extraire le contenu des documents uploadés
                docs_text = ""
                if ma_docs_upload:
                    from document_extractor import extract_text as _ext_start
                    for uf in ma_docs_upload:
                        extracted = _ext_start(uf)
                        if extracted.strip():
                            docs_text += f"\n\n--- Document fourni : {uf.name} ---\n{extracted}"
                st.session_state["ma_context_docs"] = docs_text
                # Reset wizard state for fresh start
                for k in list(st.session_state.keys()):
                    if k.startswith("q_ma_buy_wizard") or k.startswith("q_idx_ma_buy_wizard"):
                        del st.session_state[k]
                st.session_state.ma_universe    = "buy"
                st.session_state.ma_company     = buy_company.strip()
                st.session_state.ma_sector      = ""
                st.session_state.ma_step_result = {}
                st.session_state.screen         = 4
                st.rerun()

    # ── TAB 1 : Analyses rapides ───────────────────────────────────────────
    with tab_standard:
        # Company name input
        company_input = st.text_input(
            "Entreprise",
            placeholder="Nom de l'entreprise...",
            label_visibility="collapsed",
            key="company_input_field",
        )

        st.markdown("""
        <div style="
            background: linear-gradient(135deg, #111111 0%, #1A1A1A 100%);
            border-radius: 14px;
            padding: 14px 22px;
            margin: 18px 0 14px 0;
            display: flex;
            align-items: center;
            gap: 12px;
        ">
            <span style="font-size:1.3rem;">📊</span>
            <div>
                <div style="color:#FFFFFF; font-size:1.05rem; font-weight:700; letter-spacing:-0.3px;">
                    Choisissez le type d'analyse
                </div>
                <div style="color:#9CA3AF; font-size:0.82rem; margin-top:2px;">
                    Cliquez sur un livrable pour lancer l'analyse
                </div>
            </div>
        </div>
        """, unsafe_allow_html=True)

        # Couleurs d'accent par livrable
        CARD_COLORS = {
            "complet":   ("#333333", "#F5F5F5"),
            "fiche":     ("#555555", "#F5F5F5"),
            "benchmark": ("#A5614A", "#F8EEE8"),
            "manda":     ("#111111", "#F5F5F5"),
            "geo":       ("#A5944A", "#F8F3E8"),
        }

        # 2x2 + 1 grid for deliverable cards
        col_a, col_b = st.columns(2)
        col_c, col_d = st.columns(2)
        col_e, _, _ = st.columns(3)

        card_cols = [col_a, col_b, col_c, col_d, col_e]

        # Zone d'erreur visible AU-DESSUS des cartes
        error_zone = st.empty()

        clicked_key = None
        for idx, deliv in enumerate(DELIVERABLES):
            accent, bg = CARD_COLORS[deliv["key"]]
            with card_cols[idx]:
                st.markdown(f"""
                <div style="
                    background:{bg};
                    border:2px solid {accent}40;
                    border-left:4px solid {accent};
                    border-radius:12px;
                    padding:16px 18px;
                    margin-bottom:8px;
                ">
                    <span style="font-size:1.5rem;">{deliv['icon']}</span>
                    <div style="font-weight:700;color:#111111;font-size:0.95rem;margin:6px 0 3px 0;">{deliv['title']}</div>
                    <div style="font-size:0.78rem;color:#6B7280;line-height:1.4;">{deliv['desc']}</div>
                </div>
                """, unsafe_allow_html=True)
                if st.button("Sélectionner", key=f"card_{deliv['key']}", use_container_width=True):
                    clicked_key = deliv["key"]

        # Validation centralisée — messages visibles
        if clicked_key:
            if not company_input.strip():
                error_zone.warning("Veuillez entrer le nom d'une entreprise avant de choisir un type d'analyse.")
            elif not anthropic_key or not tavily_key:
                error_zone.error("Clés API manquantes. Contactez l'administrateur.")
            else:
                st.session_state.company          = company_input.strip()
                st.session_state.deliverable_type = clicked_key
                st.session_state.screen           = 2
                st.session_state.steps_done       = []
                st.session_state.current_step     = ""
                st.session_state.result_text      = ""
                st.rerun()

        # Context section — tabs for manual input and document import
        with st.expander("💡 Informations déjà connues (optionnel)"):
            tab_manual, tab_docs = st.tabs(["✏️ Saisie libre", "📎 Importer des documents"])

            with tab_manual:
                manual_context = st.text_area(
                    "Ce que vous savez déjà sur cette entreprise",
                    placeholder=(
                        "Exemples : secteur d'activité, chiffre d'affaires approximatif, "
                        "principaux clients, zone géographique, contexte de l'opération..."
                    ),
                    height=130,
                    key="context_input_field",
                )

            with tab_docs:
                st.markdown(
                    '<div class="section-label">Glissez jusqu\'à 5 fichiers (PDF, Word, Excel, TXT, CSV, MD)</div>',
                    unsafe_allow_html=True,
                )
                uploaded_files = st.file_uploader(
                    "Importer des documents",
                    type=["pdf", "docx", "xlsx", "xls", "txt", "md", "csv"],
                    accept_multiple_files=True,
                    label_visibility="collapsed",
                    key="doc_uploader",
                )

                doc_texts = []
                if uploaded_files:
                    from document_extractor import extract_text
                    shown = uploaded_files[:5]
                    for uf in shown:
                        extracted = extract_text(uf)
                        char_count = len(extracted)
                        doc_texts.append(
                            f"\n\n--- Document : {uf.name} ---\n\n{extracted}"
                        )
                        st.markdown(
                            f'<div class="file-badge">'
                            f'📄 <span class="file-badge-name">{uf.name}</span>'
                            f' &nbsp;·&nbsp; {char_count:,} car.'
                            f'</div>',
                            unsafe_allow_html=True,
                        )
                    if len(uploaded_files) > 5:
                        st.caption("Seuls les 5 premiers fichiers sont pris en compte.")

            # Combine manual text + document texts into session context
            combined_context = manual_context
            if doc_texts:
                combined_context = combined_context + "".join(doc_texts)
            st.session_state.context = combined_context


# ══════════════════════════════════════════════════════════════════════════════
# SCREEN 2 — WAITING / RUNNING
# ══════════════════════════════════════════════════════════════════════════════
elif st.session_state.screen == 2:

    company    = st.session_state.company
    deliv_key  = st.session_state.deliverable_type
    deliv_info = DELIVERABLE_BY_KEY.get(deliv_key, DELIVERABLES[0])
    context    = st.session_state.context
    all_steps  = deliv_info["steps"]

    # Animated progress bar at top
    st.markdown(
        '<div class="progress-topbar"><div class="progress-topbar-fill"></div></div>',
        unsafe_allow_html=True,
    )

    # Header with company + deliverable type
    st.markdown(
        f'<div class="company-badge">🏢 🏢 {company}</div>',
        unsafe_allow_html=True,
    )
    st.markdown(
        f'<span class="progress-deliverable">{deliv_info["icon"]} {deliv_info["title"]}</span>',
        unsafe_allow_html=True,
    )

    # Progress display placeholder
    progress_placeholder = st.empty()

    estimated = ESTIMATED_SECONDS.get(deliv_key, 180)

    def render_progress(current_step: str, steps_done: list, elapsed: float = 0):
        total  = len(all_steps)
        done_n = len(steps_done)

        # ── Timing display ──
        if elapsed > 0:
            # Temps écoulé
            m_el, s_el  = divmod(int(elapsed), 60)
            elapsed_str = f"{m_el}:{s_el:02d}"
            # Temps restant en MM:SS précis
            remaining   = max(0, estimated - elapsed)
            m_re, s_re  = divmod(int(remaining), 60)
            if remaining <= 0:
                rem_str   = "Finalisation..."
                rem_color = "#333333"
            else:
                rem_str   = f"{m_re}:{s_re:02d}"
                rem_color = "#333333" if remaining > 60 else "#E67E22"
            pct_bar = min(97, int(elapsed / estimated * 100)) if estimated else 50
            timing_html = f"""
            <div style="background:#F5F5F5; border-radius:8px; padding:12px 18px; margin-bottom:12px; font-size:0.85rem;">
                <div style="display:flex; justify-content:space-between; align-items:center; margin-bottom:8px;">
                    <span style="color:#555555;">⏱ Temps écoulé</span>
                    <span style="color:#555555;">Temps restant</span>
                </div>
                <div style="display:flex; justify-content:space-between; align-items:center;">
                    <strong style="color:#111111; font-size:1.3rem; font-family:monospace; letter-spacing:1px;">{elapsed_str}</strong>
                    <strong style="color:{rem_color}; font-size:1.3rem; font-family:monospace; letter-spacing:1px;">{rem_str}</strong>
                </div>
            </div>
            <div style="background:#E5E5E5; border-radius:4px; height:6px; margin-bottom:18px; overflow:hidden;">
                <div style="background:#111111; width:{pct_bar}%; height:100%; border-radius:4px; transition:width 0.5s;"></div>
            </div>"""
        else:
            est_min = max(1, int(estimated / 60))
            m_est, s_est = divmod(estimated, 60)
            timing_html = f"""
            <div style="background:#F5F5F5; border-radius:8px; padding:12px 18px; margin-bottom:12px;
                        display:flex; align-items:center; gap:8px; font-size:0.83rem; color:#555555;">
                <span>⏱</span>
                <span>Durée estimée : <strong style="color:#333333; font-family:monospace;">{est_min}:{s_est:02d}</strong></span>
            </div>"""

        # ── Step rows ──
        html_rows = ""
        for step in all_steps:
            if step in steps_done:
                css   = "done"
                icon  = "·"
                pulse = ""
            elif step == current_step:
                css   = "active"
                icon  = "›"
                pulse = '<span class="pulse-dot"></span>'
            else:
                css   = "pending"
                icon  = "⬜"
                pulse = ""
            html_rows += (
                f'<div class="step-row {css}">'
                f'<span class="step-icon">{icon}</span>'
                f'<span>{step}{pulse}</span>'
                f"</div>"
            )

        progress_placeholder.markdown(
            f"""
            <div class="progress-container">
                <div class="progress-header">
                    <div>
                        <div class="progress-title">Progression de l'analyse</div>
                        <div class="progress-company">{company} &nbsp;·&nbsp; {done_n}/{total} étapes</div>
                    </div>
                </div>
                {timing_html}
                {html_rows}
                <div class="searching-label">
                    <span class="pulse-dot"></span>
                    L'agent effectue des recherches web en temps réel...
                </div>
            </div>
            """,
            unsafe_allow_html=True,
        )

    render_progress("", [], elapsed=0)

    # Vérification des clés avant de lancer
    if not anthropic_key or not tavily_key:
        st.error("🔑 Clés API manquantes — impossible de lancer l'analyse. Contactez l'administrateur.")
        if st.button("← Retour"):
            st.session_state.screen = 1
            st.rerun()
        st.stop()

    # ── Run the agent ──────────────────────────────────────────────────────
    os.environ["ANTHROPIC_API_KEY"] = anthropic_key
    os.environ["TAVILY_API_KEY"]    = tavily_key

    from agent import run_screening

    full_text  = [""]
    start_time = time.time()

    def on_text(text):
        full_text[0] = text

    def on_tool_use(name, inputs):
        # Rafraîchit le timer à chaque recherche web (~toutes les 5-10s)
        elapsed = time.time() - start_time
        render_progress(st.session_state.current_step, st.session_state.steps_done, elapsed=elapsed)

    def on_tool_result(_):
        pass

    def on_step(step_name: str):
        # Mark previous current step as done
        if (
            st.session_state.current_step
            and st.session_state.current_step not in st.session_state.steps_done
        ):
            st.session_state.steps_done = st.session_state.steps_done + [st.session_state.current_step]
        st.session_state.current_step = step_name
        elapsed = time.time() - start_time
        render_progress(step_name, st.session_state.steps_done, elapsed=elapsed)

    try:
        result = run_screening(
            company_name=company,
            deliverable_type=deliv_key,
            context=context,
            on_text=on_text,
            on_tool_use=on_tool_use,
            on_tool_result=on_tool_result,
            on_step=on_step,
        )
        # Mark last step as done
        final_steps_done = list(st.session_state.steps_done)
        if st.session_state.current_step and st.session_state.current_step not in final_steps_done:
            final_steps_done.append(st.session_state.current_step)

        st.session_state.result_text = result
        st.session_state.steps_done  = final_steps_done
        st.session_state.screen      = 3
        st.rerun()
    except Exception as e:
        st.error(f"Erreur lors de l'analyse : {e}")
        if st.button("← Retour"):
            st.session_state.screen = 1
            st.rerun()


# ══════════════════════════════════════════════════════════════════════════════
# SCREEN 3 — RESULTS
# ══════════════════════════════════════════════════════════════════════════════
elif st.session_state.screen == 3:

    company    = st.session_state.company
    deliv_key  = st.session_state.deliverable_type
    deliv_info = DELIVERABLE_BY_KEY.get(deliv_key, DELIVERABLES[0])
    result     = st.session_state.result_text

    # Header
    st.markdown(
        f'<div class="company-badge">🏢 {company}</div>',
        unsafe_allow_html=True,
    )
    st.markdown(f"**{deliv_info['title']}** — Analyse terminée")
    st.markdown("---")

    # Download buttons
    col_dl1, col_dl2, col_dl3 = st.columns([2, 2, 4])

    with col_dl1:
        try:
            from word_generator import generate_word
            docx_bytes = generate_word(result, company, deliv_key)
            st.download_button(
                label="📝 Télécharger en Word",
                data=docx_bytes,
                file_name=f"screening_{company.replace(' ', '_').lower()}.docx",
                mime="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
                use_container_width=True,
            )
        except Exception as ex:
            st.warning(f"Word indisponible : {ex}")

    with col_dl2:
        try:
            from pdf_generator import generate_pdf
            pdf_bytes = generate_pdf(result, company, deliv_key)
            st.download_button(
                label="📕 Télécharger en PDF",
                data=pdf_bytes,
                file_name=f"screening_{company.replace(' ', '_').lower()}.pdf",
                mime="application/pdf",
                use_container_width=True,
            )
        except ImportError:
            st.info("PDF indisponible (reportlab non installé).")
        except Exception as ex:
            st.warning(f"PDF temporairement indisponible : {ex}")

    st.markdown("---")

    # Full report in white box
    st.markdown('<div class="result-header">Rapport d\'analyse</div>', unsafe_allow_html=True)
    with st.container(border=True):
        st.markdown(result)

    st.markdown("---")

    # ── Email de prise de contact ──────────────────────────────────────────
    import re as _re
    import anthropic as _anthropic

    def find_executives(competitor: str) -> list:
        """Cherche les dirigeants via plusieurs sources puis extrait les noms avec Claude."""
        from tavily import TavilyClient
        import anthropic as _ant
        import json as _json
        import urllib.parse as _up

        tc = TavilyClient(api_key=tavily_key)
        all_content = []

        queries = [
            f'"{competitor}" gérant président directeur général site:pappers.fr',
            f'"{competitor}" dirigeants fondateur PDG CEO direction site:societe.com OR site:verif.com',
            f'"{competitor}" équipe direction dirigeants LinkedIn',
        ]
        for q in queries:
            try:
                res = tc.search(query=q, max_results=4, search_depth="advanced")
                for r in res.get("results", []):
                    snippet = f"[{r.get('url','')}]\n{r.get('title','')}\n{r.get('content','')}"
                    all_content.append(snippet)
            except Exception:
                pass

        if not all_content:
            return []

        # Claude extrait les dirigeants intelligemment depuis les sources brutes
        combined = "\n\n---\n\n".join(all_content[:8])[:4000]
        client = _ant.Anthropic(api_key=anthropic_key)
        try:
            resp = client.messages.create(
                model="claude-haiku-4-5",
                max_tokens=400,
                messages=[{"role": "user", "content":
                    f"""Extrais les noms et titres des dirigeants actuels de "{competitor}" depuis ces résultats.

{combined}

Règles :
- Uniquement des personnes en poste actuellement (pas d'anciens dirigeants)
- Titres acceptés : PDG, DG, Président, Directeur, Fondateur, Gérant, CEO, CFO, COO, VP, Associé, Partner
- Maximum 5 personnes
- Si vraiment aucun dirigeant identifiable, retourne []

Retourne UNIQUEMENT ce JSON (rien d'autre) :
[{{"name": "Prénom Nom", "title": "Titre"}}]"""}]
            )
            text = resp.content[0].text.strip()
            match = _re.search(r'\[.*?\]', text, _re.DOTALL)
            if match:
                raw = _json.loads(match.group())
                executives = []
                for e in raw:
                    n, t = e.get("name","").strip(), e.get("title","").strip()
                    if n and t and len(n) > 3:
                        executives.append({"name": n, "title": t})
                # Cherche l'URL directe du profil LinkedIn pour chaque dirigeant
                comp_lower = competitor.lower()
                for ex in executives[:5]:
                    profile_url = None
                    try:
                        li_res = tc.search(
                            query=f'"{ex["name"]}" "{competitor}" site:linkedin.com/in',
                            max_results=5,
                            search_depth="basic",
                        )
                        for r in li_res.get("results", []):
                            u = r.get("url", "")
                            if "linkedin.com/in/" not in u:
                                continue
                            # Vérifie que le snippet mentionne bien l'entreprise
                            snippet = (r.get("content", "") + r.get("title", "")).lower()
                            if comp_lower in snippet or ex["name"].split()[0].lower() in snippet:
                                profile_url = u.split("?")[0].rstrip("/")
                                break
                    except Exception:
                        pass
                    # Fallback : recherche LinkedIn avec nom + société (résultats filtrés)
                    if not profile_url:
                        kw = _up.quote(f"{ex['name']} {competitor}")
                        profile_url = f"https://www.linkedin.com/search/results/people/?keywords={kw}"
                    ex["url"] = profile_url
                return executives[:5]
        except Exception:
            pass
        return []

    def _extract_competitor_context(competitor: str, report: str) -> str:
        """Extrait la fiche du concurrent depuis le rapport."""
        # Cherche la section ### ... NOM jusqu'au prochain --- ou ###
        pattern = rf'###[^\n]*{_re.escape(competitor)}[^\n]*\n(.*?)(?=\n---|\n###|\Z)'
        match = _re.search(pattern, report, _re.DOTALL | _re.IGNORECASE)
        if match:
            raw = match.group(1).strip()
            # Garde max 800 caractères pour ne pas surcharger le prompt
            return raw[:800]
        return ""

    EMAIL_TYPE_LABELS = {
        "rachat":   "Investisseur — Rachat",
        "levee":    "M&A — Levée de fonds",
        "buildup":  "Investisseur — Build-up",
    }

    EMAIL_TYPE_PROMPTS = {
        "rachat": """Tu es un investisseur ou un fonds qui envisage d'acquérir {competitor}.
Tu as fait tes recherches sur leur secteur et leur activité de façon indépendante — ne mentionne aucune autre société ni aucune source dans le message.
{context_block}
Rédige un message LinkedIn de premier contact, discret et professionnel.

CONTRAINTES ABSOLUES :
- Ne mentionne jamais le nom d'une autre entreprise, ni pourquoi tu les as repérés
- Objet (max 200 caractères) : sobre, sans les mots "rachat" ou "acquisition"
- Corps du message : 900 caractères MAXIMUM (LinkedIn InMail) — sois très concis
- Commence par "Bonjour [Prénom]," — laisse "[Prénom]" entre crochets, n'invente jamais un prénom
- Ligne 1 : qui tu es en une phrase (fonds/investisseur, secteur d'activité)
- Ligne 2 : 1 observation concrète sur {competitor} qui montre que tu les connais — sans citer de source
- Ligne 3 : proposer un échange de 20 min, confidentiel, sans engagement
- Ton : direct, humain, jamais corporatif — aucun jargon (pas de "synergies", "due diligence", "deal flow")
- Pas de mise en forme : pas de tirets, pas de gras, pas de listes
- Signature : "[Prénom Nom] — [Titre], [Fonds]"
- Entièrement en français""",

        "levee": """Tu es un conseiller M&A accompagnant des dirigeants dans des opérations de financement et de croissance.
Tu as identifié {competitor} comme une société à fort potentiel — de façon indépendante, sans mentionner aucune autre société ni source dans le message.
{context_block}
Rédige un message LinkedIn de premier contact, orienté opportunité de croissance.

CONTRAINTES ABSOLUES :
- Ne mentionne jamais le nom d'une autre entreprise, ni la démarche qui t'a conduit à les contacter
- Objet (max 200 caractères) : axé croissance, sans les mots "levée de fonds" ou "financement"
- Corps du message : 900 caractères MAXIMUM (LinkedIn InMail) — sois très concis
- Commence par "Bonjour [Prénom]," — laisse "[Prénom]" entre crochets, n'invente jamais un prénom
- Ligne 1 : qui tu es en une phrase (conseil M&A, réseau de fonds, secteur)
- Ligne 2 : 1 observation sur leur marché ou leur activité qui justifie l'intérêt — sans citer de source
- Ligne 3 : proposer un échange de 20 min pour voir si tu peux les aider à accélérer
- Ton : conseiller, pas vendeur — facilitateur qui apporte de la valeur
- Aucun anglicisme excessif, aucun jargon corporate
- Pas de mise en forme : pas de tirets, pas de gras, pas de listes
- Signature : "[Prénom Nom] — [Titre], [Cabinet]"
- Entièrement en français""",

        "buildup": """Tu es un investisseur actionnaire d'une société active dans le même secteur que {competitor}, avec une stratégie de consolidation — tu cherches des acteurs complémentaires pour construire un leader de marché.
Tu as identifié {competitor} de façon indépendante — ne mentionne aucune autre société ni aucune source dans le message.
{context_block}
Rédige un message LinkedIn de premier contact, entre pairs du même secteur.

CONTRAINTES ABSOLUES :
- Ne mentionne jamais le nom d'une autre société, ni la démarche qui t'a mené à les contacter
- Objet (max 200 caractères) : axé vision commune du marché, sans les mots "build-up", "rachat" ou "acquisition"
- Corps du message : 900 caractères MAXIMUM (LinkedIn InMail) — sois très concis
- Commence par "Bonjour [Prénom]," — laisse "[Prénom]" entre crochets, n'invente jamais un prénom
- Ligne 1 : qui tu es, ta société et son positionnement dans ce secteur (une phrase)
- Ligne 2 : 1 point de complémentarité ou de vision partagée avec {competitor} — sans citer de source
- Ligne 3 : proposer un échange informel de 20 min pour partager les perspectives du marché
- Ton : entrepreneur à entrepreneur, long terme, pas financier
- Zéro jargon : pas de "build-up", "synergies", "closing", "multiple", "plateforme"
- Pas de mise en forme : pas de tirets, pas de gras, pas de listes
- Signature : "[Prénom Nom] — [Titre], [Société]"
- Entièrement en français""",
    }

    def generate_contact_email(competitor: str, main_company: str, comp_context: str = "", email_type: str = "rachat") -> str:
        client = _anthropic.Anthropic(api_key=anthropic_key)

        context_block = ""
        if comp_context:
            context_block = (
                f"\nVoici ce que tu sais sur {competitor} (utilise 1 ou 2 éléments concrets) :\n"
                f"---\n{comp_context}\n---\n"
            )

        prompt_template = EMAIL_TYPE_PROMPTS.get(email_type, EMAIL_TYPE_PROMPTS["rachat"])
        prompt = prompt_template.format(
            competitor=competitor,
            main_company=main_company,
            context_block=context_block,
        )

        response = client.messages.create(
            model="claude-sonnet-4-6",
            max_tokens=700,
            messages=[{"role": "user", "content":
                f"{prompt}\n\nRetourne uniquement l'email, format exact :\nObjet : [objet]\n\n[corps]"}]
        )
        return response.content[0].text

    import urllib.parse as _urlparse

    li_svg = ('<svg xmlns="http://www.w3.org/2000/svg" width="14" height="14" '
              'viewBox="0 0 24 24" fill="#555555"><path d="M20.447 20.452h-3.554v-5.569'
              'c0-1.328-.027-3.037-1.852-3.037-1.853 0-2.136 1.445-2.136 2.939v5.667H9.351'
              'V9h3.414v1.561h.046c.477-.9 1.637-1.85 3.37-1.85 3.601 0 4.267 2.37 4.267 '
              '5.455v6.286zM5.337 7.433a2.062 2.062 0 0 1-2.063-2.065 2.064 2.064 0 1 1 '
              '2.063 2.065zm1.782 13.019H3.555V9h3.564v11.452zM22.225 0H1.771C.792 0 0 '
              '.774 0 1.729v20.542C0 23.227.792 24 1.771 24h20.451C23.2 24 24 23.227 24 '
              '22.271V1.729C24 .774 23.2 0 22.222 0h.003z"/></svg>')

    # ══════════════════════════════════════════════════════════════════════
    # BENCHMARK : email de prise de contact par concurrent
    # ══════════════════════════════════════════════════════════════════════
    if deliv_key == "benchmark":

        # Détection stricte des concurrents (patterns benchmark uniquement)
        detected = _re.findall(r'###\s+Concurrent[^—\n]*—\s*\**([^\n*]+)', result)
        detected += _re.findall(r'###\s+\d+[^\S\n]*[—–-][^\S\n]*\**([^\n*]+)', result)
        seen, unique = set(), []
        for c in detected:
            c = c.strip().strip("*").strip()
            if c and c not in seen:
                seen.add(c)
                unique.append(c)
        detected = unique

        st.markdown("""
        <div style="background:#111111;
                    border-radius:14px; padding:20px 24px; margin-bottom:20px;">
            <div style="color:#FFFFFF; font-size:1.05rem; font-weight:700; margin-bottom:4px;">
                Préparer une prise de contact
            </div>
            <div style="color:#9CA3AF; font-size:0.83rem;">
                Choisissez le type d'approche, puis cliquez sur un concurrent pour générer le message LinkedIn.
            </div>
        </div>
        """, unsafe_allow_html=True)

        # Sélecteur type d'email
        _etype_current = st.session_state.get("email_type", "rachat")
        _etype_options = [
            ("rachat",  "Investisseur — Rachat"),
            ("levee",   "M&A — Levée de fonds"),
            ("buildup", "Investisseur — Build-up"),
        ]
        _etype_cols = st.columns(3)
        for _ec, (_ek, _el) in zip(_etype_cols, _etype_options):
            _css_class = "etype-btn-selected" if _etype_current == _ek else "etype-btn-idle"
            _ec.markdown(f'<div class="{_css_class}">', unsafe_allow_html=True)
            with _ec:
                if st.button(_el, key=f"etype_{_ek}", use_container_width=True):
                    st.session_state.email_type = _ek
                    st.rerun()
            _ec.markdown('</div>', unsafe_allow_html=True)

        st.markdown('<div style="margin-top:12px;"></div>', unsafe_allow_html=True)

        if detected:
            cols_per_row = 3
            rows = [detected[i:i+cols_per_row] for i in range(0, len(detected), cols_per_row)]
            for row in rows:
                cols = st.columns(len(row))
                for col, comp in zip(cols, row):
                    with col:
                        if st.button(f"{comp}", key=f"email_btn_{comp}", use_container_width=True):
                            st.session_state.email_target     = comp
                            st.session_state.generated_email  = ""
                            st.session_state.found_executives = []
                            with st.spinner("Rédaction du message et recherche des dirigeants..."):
                                ctx = _extract_competitor_context(comp, result)
                                st.session_state.generated_email  = generate_contact_email(comp, company, ctx, email_type=st.session_state.email_type)
                                st.session_state.found_executives = find_executives(comp)
        else:
            manual_comp = st.text_input("Nom du concurrent à contacter :",
                                         placeholder="ex : Entreprise XYZ",
                                         key="manual_email_target")
            if manual_comp and st.button("✉️ Générer le message", type="primary"):
                st.session_state.email_target     = manual_comp
                st.session_state.generated_email  = ""
                st.session_state.found_executives = []
                with st.spinner("Rédaction du message et recherche des dirigeants..."):
                    ctx = _extract_competitor_context(manual_comp, result)
                    st.session_state.generated_email  = generate_contact_email(manual_comp, company, ctx, email_type=st.session_state.email_type)
                    st.session_state.found_executives = find_executives(manual_comp)

        # Afficher le message généré
        if st.session_state.generated_email:
            raw = st.session_state.generated_email.strip()
            lines = raw.split("\n")
            subject_line, body_lines, in_body = "", [], False
            for line in lines:
                if line.lower().startswith("objet"):
                    subject_line = line.split(":", 1)[-1].strip()
                elif in_body or (subject_line and line.strip() == "" and not in_body):
                    in_body = True
                    body_lines.append(line)
                elif subject_line and line.strip():
                    in_body = True
                    body_lines.append(line)
            body_text = "\n".join(body_lines).strip()

            st.markdown(
                f'<div style="font-size:0.82rem;color:#6B7280;margin:18px 0 10px 0;">'
                f'Brouillon pour <strong style="color:#111111;">{st.session_state.email_target}</strong>'
                f' — personnalisez les champs entre crochets avant envoi</div>',
                unsafe_allow_html=True,
            )
            st.markdown(f"""
<div style="background:#FFFFFF; border:1px solid #E5E5E5; border-radius:12px;
            padding:0; overflow:hidden; box-shadow:0 2px 10px rgba(0,0,0,0.06);">
  <div style="background:#F5F5F5; border-bottom:1px solid #E5E5E5;
              padding:12px 20px; display:flex; align-items:center; gap:10px;">
    <span style="font-size:0.78rem; font-weight:600; color:#64748B; min-width:48px;">Objet</span>
    <span style="font-size:0.92rem; font-weight:600; color:#111111;">{subject_line}</span>
  </div>
  <div style="padding:22px 24px; font-size:0.93rem; color:#1A202C;
              line-height:1.85; white-space:pre-wrap; font-family:'Inter', sans-serif;">
{body_text}
  </div>
</div>
""", unsafe_allow_html=True)

            with st.expander("📋 Copier le texte brut"):
                st.text_area("", value=raw, height=220, label_visibility="collapsed", key="email_copy_area")

            li_company = f"https://www.linkedin.com/search/results/companies/?keywords={_urlparse.quote(st.session_state.email_target)}"
            st.markdown(
                f'<div style="margin-top:10px;">'
                f'<a href="{li_company}" target="_blank" style="display:inline-flex;align-items:center;'
                f'gap:6px;background:#FFFFFF;border:1px solid #E5E5E5;border-radius:8px;'
                f'padding:7px 14px;font-size:0.83rem;font-weight:600;color:#555555;text-decoration:none;">'
                f'{li_svg} Voir la page entreprise</a></div>',
                unsafe_allow_html=True,
            )

            execs = st.session_state.get("found_executives", [])
            if execs:
                st.markdown(
                    '<div style="font-size:0.82rem;font-weight:600;color:#64748B;'
                    'text-transform:uppercase;letter-spacing:0.05em;margin:18px 0 10px 0;">'
                    'Dirigeants identifiés'
                    '<span style="font-size:0.72rem;font-weight:400;color:#9CA3AF;'
                    'text-transform:none;margin-left:6px;">(Pappers / Infogreffe)</span></div>',
                    unsafe_allow_html=True,
                )
                cols = st.columns(min(len(execs), 3))
                for idx, exec_ in enumerate(execs):
                    with cols[idx % 3]:
                        st.markdown(
                            f'<a href="{exec_["url"]}" target="_blank" style="text-decoration:none;">'
                            f'<div style="background:#FFFFFF;border:1px solid #E5E5E5;border-radius:10px;'
                            f'padding:12px 14px;">'
                            f'<div style="display:flex;align-items:center;gap:8px;margin-bottom:6px;">'
                            f'{li_svg}'
                            f'<span style="font-size:0.88rem;font-weight:700;color:#111111;">{exec_["name"]}</span>'
                            f'</div>'
                            f'<div style="font-size:0.78rem;color:#64748B;">{exec_["title"] or "Dirigeant"}</div>'
                            f'<div style="font-size:0.75rem;color:#555555;margin-top:6px;font-weight:600;">Voir le profil →</div>'
                            f'</div></a>',
                            unsafe_allow_html=True,
                        )
            else:
                st.markdown(
                    '<div style="font-size:0.8rem;color:#9CA3AF;margin-top:12px;font-style:italic;">'
                    'Aucun profil dirigeant trouvé — utilisez la page entreprise LinkedIn.</div>',
                    unsafe_allow_html=True,
                )

    # ══════════════════════════════════════════════════════════════════════
    # AUTRES LIVRABLES : liens LinkedIn pertinents extraits du rapport
    # ══════════════════════════════════════════════════════════════════════
    else:
        import anthropic as _ant
        import json as _json

        def extract_linkedin_entities(report: str, deliv_type: str, main_co: str) -> list:
            """Extrait via Claude Haiku les entités les plus pertinentes du rapport."""
            instructions = {
                "fiche": (
                    "Ce rapport est une fiche entreprise. "
                    "Retourne la société principale et ses filiales ou partenaires clés mentionnés."
                ),
                "manda": (
                    "Ce rapport contient des actualités M&A. "
                    "Extrais les acquéreurs actifs, fonds d'investissement et fédérations professionnelles mentionnés."
                ),
                "geo": (
                    "Ce rapport est une analyse géographique. "
                    "Extrais les entreprises et acteurs locaux les plus importants mentionnés par zone."
                ),
            }
            instruction = instructions.get(deliv_type, "Extrais les principales organisations mentionnées.")
            client = _ant.Anthropic(api_key=anthropic_key)
            try:
                resp = client.messages.create(
                    model="claude-haiku-4-5",
                    max_tokens=350,
                    messages=[{"role": "user", "content":
                        f"""{instruction}

{report[:3500]}

Règles :
- Maximum 6 entités, pertinentes pour une prise de contact professionnelle
- Exclure "{main_co}" lui-même
- Retourne UNIQUEMENT ce JSON :
[{{"name": "Nom exact", "type": "Acquéreur / Fonds / Fédération / Entreprise"}}]"""}]
                )
                text = resp.content[0].text.strip()
                match = _re.search(r'\[.*?\]', text, _re.DOTALL)
                if match:
                    raw = _json.loads(match.group())
                    entities = []
                    for e in raw:
                        n = e.get("name", "").strip()
                        t = e.get("type", "").strip()
                        if n and len(n) > 2:
                            li_url = f"https://www.linkedin.com/search/results/companies/?keywords={_urlparse.quote(n)}"
                            entities.append({"name": n, "type": t, "url": li_url})
                    return entities[:6]
            except Exception:
                pass
            return []

        st.markdown("""
        <div style="background:#111111;
                    border-radius:14px; padding:20px 24px; margin-bottom:20px;">
            <div style="color:#FFFFFF; font-size:1.05rem; font-weight:700; margin-bottom:4px;">
                Liens LinkedIn pertinents
            </div>
            <div style="color:#9CA3AF; font-size:0.83rem;">
                Acteurs clés identifiés dans ce rapport — cliquez pour accéder directement à leur page.
            </div>
        </div>
        """, unsafe_allow_html=True)

        li_key = f"li_entities_{deliv_key}_{company}"
        if li_key not in st.session_state:
            with st.spinner("Identification des contacts LinkedIn pertinents..."):
                st.session_state[li_key] = extract_linkedin_entities(result, deliv_key, company)

        entities = st.session_state.get(li_key, [])
        if entities:
            cols_per_row = 3
            rows = [entities[i:i+cols_per_row] for i in range(0, len(entities), cols_per_row)]
            for row in rows:
                cols = st.columns(len(row))
                for col, ent in zip(cols, row):
                    with col:
                        st.markdown(
                            f'<a href="{ent["url"]}" target="_blank" style="text-decoration:none;">'
                            f'<div style="background:#FFFFFF;border:1px solid #E5E5E5;border-radius:10px;'
                            f'padding:14px 16px;margin-bottom:8px;">'
                            f'<div style="display:flex;align-items:center;gap:8px;margin-bottom:6px;">'
                            f'{li_svg}'
                            f'<span style="font-size:0.88rem;font-weight:700;color:#111111;">{ent["name"]}</span>'
                            f'</div>'
                            f'<div style="font-size:0.75rem;color:#64748B;">{ent["type"]}</div>'
                            f'<div style="font-size:0.75rem;color:#555555;margin-top:6px;font-weight:600;">Voir sur LinkedIn →</div>'
                            f'</div></a>',
                            unsafe_allow_html=True,
                        )
        else:
            st.markdown(
                '<div style="font-size:0.83rem;color:#9CA3AF;font-style:italic;">'
                'Aucune entité identifiée dans ce rapport.</div>',
                unsafe_allow_html=True,
            )

    st.markdown("---")

    # New analysis button
    if st.button("🔍 Nouvelle analyse", type="primary"):
        for k in ["screen", "company", "deliverable_type", "context", "result_text",
                  "current_step", "steps_done", "email_target", "generated_email",
                  "found_executives", "email_type"]:
            if k == "screen":
                st.session_state[k] = 1
            elif k in ("steps_done", "found_executives"):
                st.session_state[k] = []
            elif k == "email_type":
                st.session_state[k] = "rachat"
            else:
                st.session_state[k] = ""
        st.rerun()

    # Disclaimer légal
    st.markdown(
        """
        <div style="
            margin-top: 32px;
            padding: 14px 18px;
            background: #F0EDE8;
            border-left: 3px solid #C8B89A;
            border-radius: 6px;
            font-size: 0.76rem;
            color: #7A6E62;
            line-height: 1.6;
        ">
        <strong style="color:#555555;">⚠️ Avertissement</strong> — Les informations présentées dans ce rapport sont
        issues de sources publiques disponibles sur Internet à la date de l'analyse. Elles sont fournies
        à titre indicatif et ne constituent en aucun cas un conseil en investissement, une recommandation
        financière ou une due diligence. Les données (chiffre d'affaires, effectifs, valorisations) sont
        des estimations et doivent être vérifiées auprès des sources primaires avant toute prise de décision.
        </div>
        """,
        unsafe_allow_html=True,
    )



# ══════════════════════════════════════════════════════════════════════════════
# SCREEN 4 — MISSION M&A
# ══════════════════════════════════════════════════════════════════════════════
elif st.session_state.screen == 4:
    from ma_agent import run_ma_module
    import re as _re4

    ma_company   = st.session_state.get("ma_company", "")
    company_info = st.session_state.get("ma_company_info", None)
    ma_phase     = st.session_state.get("ma_phase", "research")

    _KNOWN_PHASES = {"wizard_pos", "run_v", "check_v", "run_h", "check_h",
                     "wizard_cats", "run_cibles", "check_cibles", "done"}
    if st.session_state.get("ma_wizard_confirmed") and ma_phase not in _KNOWN_PHASES:
        st.session_state.ma_phase = "wizard_pos"
        ma_phase = "wizard_pos"

    os.environ["ANTHROPIC_API_KEY"] = anthropic_key
    os.environ["TAVILY_API_KEY"]    = tavily_key

    def _s4_quit():
        prefix_keys = ("q_ma_wiz", "q_idx_ma_wiz", "wc_",
                       "ma_company", "ma_wizard", "ma_phase",
                       "ma_result", "ma_cibles", "ma_variables", "ma_mission",
                       "ma_categories", "ma_current", "ma_context", "ma_step",
                       "ma_wc_")
        for k in list(st.session_state.keys()):
            if any(k.startswith(p) for p in prefix_keys):
                del st.session_state[k]
        st.session_state.screen = 1

    # ── En-tête ────────────────────────────────────────────────────────────
    hdr_col, quit_col = st.columns([7, 1])
    with hdr_col:
        st.markdown(
            f'<div style="font-size:1.05rem;font-weight:700;margin-bottom:2px;">{ma_company}</div>'
            f'<div style="font-size:0.78rem;color:#777777;margin-bottom:6px;">Screening Buy Side</div>',
            unsafe_allow_html=True,
        )
    with quit_col:
        if st.button("✕ Quitter", use_container_width=True):
            _s4_quit()
            st.rerun()

    st.markdown('<div style="height:8px"></div>', unsafe_allow_html=True)

    # ── Helpers ────────────────────────────────────────────────────────────
    def _run_module_s4(mod_key, mod_input, mod_label, next_phase, variables):
        """Lance un module, sauvegarde le résultat, bascule la phase."""
        out_ph = st.empty()
        st.markdown(
            f'<div style="border:1px solid #CCCCCC;border-radius:10px;background:#FFFFFF;'
            f'padding:14px 20px;font-size:0.88rem;font-weight:600;color:#111111;">'
            f'⏳ &nbsp;{mod_label} en cours…</div>',
            unsafe_allow_html=True,
        )
        def _on_text(text, _ph=out_ph):
            _ph.markdown(
                '<div style="border:1px solid #E0E0E0;border-radius:10px;background:#FAFAFA;'
                'padding:16px 20px;max-height:260px;overflow-y:auto;'
                'font-size:0.87rem;line-height:1.8;color:#333333;">'
                + text[:4000].replace("\n", "<br>") + '</div>',
                unsafe_allow_html=True,
            )
        try:
            result = run_ma_module(
                module_key=mod_key, company=ma_company, sector="",
                input_data=mod_input, variables=variables, on_text=_on_text,
            )
            st.session_state[f"ma_result_{mod_key}"] = result
            st.session_state.ma_phase = next_phase
            out_ph.empty()
            st.rerun()
        except Exception as _e:
            import traceback
            out_ph.empty()
            st.error(f"❌ Erreur : {_e}")
            with st.expander("Détail"):
                st.code(traceback.format_exc())

    def _satisfaction_buttons(label, yes_phase, no_phase, no_clears=None):
        """Affiche les deux boutons satisfaction — renvoie True si on a cliqué (pour stopper le rendu)."""
        st.markdown(f'<div style="font-size:0.88rem;font-weight:600;margin:12px 0 6px;">Es-tu satisfait du résultat ?</div>', unsafe_allow_html=True)
        c1, c2 = st.columns(2)
        with c1:
            if st.button(f"✓ Oui — continuer", type="primary", use_container_width=True, key=f"sat_yes_{yes_phase}"):
                st.session_state.ma_phase = yes_phase
                st.rerun()
        with c2:
            if st.button(f"↺ Non — relancer", use_container_width=True, key=f"sat_no_{no_phase}"):
                if no_clears:
                    for _k in no_clears:
                        st.session_state.pop(_k, None)
                st.session_state.ma_phase = no_phase
                st.rerun()

    def _result_card(mod_key, mod_label, result_text):
        """Carte de résultat avec métadonnées + Excel + expandeur."""
        n_rows = len(_re4.findall(r'^\|[^-\|]', result_text, _re4.MULTILINE))
        n_src  = len(_re4.findall(r'https?://', result_text))
        meta   = f"{n_rows} lignes" if n_rows > 1 else f"{len(result_text.split())} mots"
        if n_src:
            meta += f" · {n_src} sources"
        with st.container(border=True):
            st.markdown(
                f'<div style="display:flex;justify-content:space-between;align-items:center;margin-bottom:4px;">'
                f'<span style="font-size:0.7rem;font-weight:700;color:#9CA3AF;text-transform:uppercase;letter-spacing:0.08em;">{mod_label}</span>'
                f'<span style="font-size:0.72rem;color:#9CA3AF;">{meta}</span></div>',
                unsafe_allow_html=True,
            )
            xlsx = _generate_single_xlsx(mod_key, mod_label, result_text, ma_company)
            st.download_button(
                "📥 Excel",
                data=xlsx,
                file_name=f"{ma_company.replace(' ', '_')}_{mod_key}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key=f"dl_{mod_key}_{ma_phase}",
            )
            with st.expander("Voir le résultat"):
                st.markdown(result_text)

    # ── Phase : Research ────────────────────────────────────────────────────
    if company_info is None:
        with st.spinner(f"Recherche d'informations sur **{ma_company}**…"):
            try:
                info = _research_company(ma_company, anthropic_key, tavily_key)
                st.session_state.ma_company_info = info if info else {}
            except Exception:
                st.session_state.ma_company_info = {}
        st.rerun()

    # ── Phase : Validation entreprise ──────────────────────────────────────
    elif not st.session_state.get("ma_wizard_confirmed", False):
        info = company_info
        description = info.get("description", f"Entreprise identifiée : {ma_company}")
        faits = info.get("faits", [])
        with st.container(border=True):
            st.markdown(
                f'<div style="font-size:0.7rem;font-weight:700;color:#9CA3AF;text-transform:uppercase;'
                f'letter-spacing:0.09em;margin-bottom:10px;">Entreprise identifi&#233;e par l&#39;IA</div>'
                f'<div style="font-size:1.2rem;font-weight:700;color:#111111;margin-bottom:6px;">{ma_company}</div>'
                f'<div style="font-size:0.9rem;color:#444444;margin-bottom:16px;line-height:1.6;">{description}</div>',
                unsafe_allow_html=True,
            )
            if faits:
                bullets = "".join(
                    f'<div style="display:flex;gap:8px;margin-bottom:7px;">'
                    f'<span style="color:#9CA3AF;flex-shrink:0;">—</span>'
                    f'<span style="font-size:0.87rem;color:#333333;">{f_}</span></div>'
                    for f_ in faits[:5]
                )
                st.markdown(
                    f'<div style="background:#F5F5F5;border-radius:8px;padding:14px 16px;margin-bottom:16px;">{bullets}</div>',
                    unsafe_allow_html=True,
                )
            c1, c2 = st.columns(2)
            with c1:
                if st.button("✓  C'est la bonne société", type="primary", use_container_width=True):
                    st.session_state.ma_wizard_confirmed = True
                    st.session_state.ma_phase = "wizard_pos"
                    st.rerun()
            with c2:
                if st.button("✗  Ce n'est pas la bonne", use_container_width=True):
                    for k in list(st.session_state.keys()):
                        if k in ("ma_company_info", "ma_wizard_confirmed"):
                            del st.session_state[k]
                    st.session_state.screen = 1
                    st.rerun()

    # ── Phase : Wizard positionnement ──────────────────────────────────────
    elif ma_phase == "wizard_pos":
        info = company_info or {}
        pos_opts = info.get("positionnement_options",
                            ["Intégrateur industriel", "Bureau d'études", "Sous-traitant", "Fabricant d'équipements"])
        _wiz_pos = [{"key": "positionnement",
                     "label": f"Quel est le positionnement de {ma_company} ?",
                     "type": "chips_or_custom", "options": pos_opts,
                     "hint": "Décrivez le positionnement…", "required": True,
                     "multi": True}]
        _, _vars_pos, _ready_pos = _render_question_cards("ma_wiz_pos", {"questions": _wiz_pos}, ma_company)
        if _ready_pos:
            st.session_state["ma_variables"] = {"positionnement": _vars_pos.get("positionnement", "")}
            st.markdown('<div style="height:6px"></div>', unsafe_allow_html=True)
            if st.button("Lancer les mappings →", type="primary", use_container_width=True):
                st.session_state.ma_phase = "run_h"
                st.rerun()

    # ── Phase : Mapping horizontal ──────────────────────────────────────────
    elif ma_phase == "run_h":
        context_docs = st.session_state.get("ma_context_docs", "")
        variables    = st.session_state.get("ma_variables", {})
        _run_module_s4("buy_02_carto_horizontale", context_docs, "Cartographie horizontale", "check_h", variables)

    elif ma_phase == "check_h":
        result_h = st.session_state.get("ma_result_buy_02_carto_horizontale", "")
        _result_card("buy_02_carto_horizontale", "Cartographie horizontale", result_h)
        _satisfaction_buttons("h", "run_v", "run_h", ["ma_result_buy_02_carto_horizontale"])

    # ── Phase : Mapping vertical ────────────────────────────────────────────
    elif ma_phase == "run_v":
        context_docs = st.session_state.get("ma_context_docs", "")
        variables    = st.session_state.get("ma_variables", {})
        _run_module_s4("buy_01_carto_verticale", context_docs, "Cartographie verticale", "check_v", variables)

    elif ma_phase == "check_v":
        result_v = st.session_state.get("ma_result_buy_01_carto_verticale", "")
        _result_card("buy_01_carto_verticale", "Cartographie verticale", result_v)
        _satisfaction_buttons("v", "wizard_cats", "run_v", ["ma_result_buy_01_carto_verticale"])

    # ── Phase : Wizard catégories + paramètres ─────────────────────────────
    elif ma_phase == "wizard_cats":
        info = company_info or {}
        cat_opts = info.get("categories_options",
                            ["Intégration systèmes", "Maintenance industrielle", "Automatisme", "Mécanique", "Électricité industrielle"])
        context_docs = st.session_state.get("ma_context_docs", "")

        # Récap mappings déjà produits
        col_h, col_v = st.columns(2)
        with col_h:
            rh = st.session_state.get("ma_result_buy_02_carto_horizontale", "")
            if rh:
                with st.container(border=True):
                    st.markdown('<div style="font-size:0.78rem;font-weight:600;color:#065F46;">✓ Mapping horizontal</div>', unsafe_allow_html=True)
                    _xh = _generate_single_xlsx("buy_02_carto_horizontale", "Mapping horizontal", rh, ma_company)
                    st.download_button("📥 Excel", data=_xh,
                                       file_name=f"{ma_company.replace(' ', '_')}_mapping_h.xlsx",
                                       mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                       key="dl_h_recap")
        with col_v:
            rv = st.session_state.get("ma_result_buy_01_carto_verticale", "")
            if rv:
                with st.container(border=True):
                    st.markdown('<div style="font-size:0.78rem;font-weight:600;color:#065F46;">✓ Mapping vertical</div>', unsafe_allow_html=True)
                    _xv = _generate_single_xlsx("buy_01_carto_verticale", "Mapping vertical", rv, ma_company)
                    st.download_button("📥 Excel", data=_xv,
                                       file_name=f"{ma_company.replace(' ', '_')}_mapping_v.xlsx",
                                       mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                       key="dl_v_recap")

        st.markdown('<div style="height:14px"></div>', unsafe_allow_html=True)

        _wc_idx = st.session_state.get("ma_wc_idx", 0)
        _wc_vars = st.session_state.get("ma_wc_vars", {})

        _wc_questions = [
            {"key": "categories", "label": "Quelles catégories de cibles visez-vous ?",
             "type": "chips_multi", "options": cat_opts},
            {"key": "zone_geo", "label": "Zone géographique cible",
             "type": "chips_single", "options": ["France uniquement", "France + Belgique", "France + DACH", "Europe"],
             "hint": "Autre zone…"},
            {"key": "fourchette_ca", "label": "Fourchette de CA des cibles",
             "type": "chips_single", "options": ["< 5 M€", "5–20 M€", "20–50 M€", "50–100 M€"],
             "hint": "Autre fourchette…"},
            {"key": "exclusions", "label": "Entreprises à exclure (optionnel)",
             "type": "textarea", "hint": "Laissez vide si aucune. Une société par ligne."},
        ]
        _wc_n = len(_wc_questions)

        # Indicateur de progression
        _wc_dots = ""
        for _i in range(_wc_n):
            if _i < _wc_idx:
                _wc_dots += '<div style="width:8px;height:8px;border-radius:50%;background:#111111;flex-shrink:0;"></div>'
            elif _i == _wc_idx:
                _wc_dots += '<div style="width:10px;height:10px;border-radius:50%;background:#111111;flex-shrink:0;"></div>'
            else:
                _wc_dots += '<div style="width:8px;height:8px;border-radius:50%;background:#D5D5D5;flex-shrink:0;"></div>'
        st.markdown(
            f'<div style="display:flex;align-items:center;gap:7px;margin-bottom:6px;">'
            f'{_wc_dots}'
            f'<span style="font-size:0.76rem;color:#9CA3AF;margin-left:3px;">{_wc_idx + 1} / {_wc_n}</span>'
            f'</div>', unsafe_allow_html=True,
        )

        if _wc_idx < _wc_n:
            _wc_q = _wc_questions[_wc_idx]
            _wc_qk = _wc_q["key"]
            _wc_type = _wc_q["type"]

            with st.container(border=True):
                st.markdown(f'<p class="q-card-label">{_wc_q["label"]}</p>', unsafe_allow_html=True)

                _wc_value = None
                _wc_ready = False

                if _wc_type == "chips_multi":
                    _wc_chosen = st.pills(
                        "opts", _wc_q["options"] + ["Autre..."],
                        selection_mode="multi", label_visibility="collapsed",
                        key=f"wc_{_wc_qk}_pill",
                    )
                    _wc_has_autre = "Autre..." in (_wc_chosen or [])
                    _wc_real = [c for c in (_wc_chosen or []) if c != "Autre..."]
                    if _wc_has_autre:
                        _wc_custom = st.text_area(
                            "Précisez", placeholder="Une catégorie par ligne…",
                            key=f"wc_{_wc_qk}_custom", label_visibility="collapsed",
                        )
                        if _wc_custom.strip():
                            _wc_real = _wc_real + [l.strip() for l in _wc_custom.strip().split("\n") if l.strip()]
                    _wc_value = _wc_real
                    _wc_ready = bool(_wc_real)

                elif _wc_type == "chips_single":
                    _wc_chosen = st.pills(
                        "opts", _wc_q["options"] + ["Autre..."],
                        selection_mode="single", label_visibility="collapsed",
                        key=f"wc_{_wc_qk}_pill",
                    )
                    _wc_has_autre = _wc_chosen == "Autre..."
                    if _wc_has_autre:
                        _wc_value = st.text_input(
                            "Précisez", placeholder=_wc_q.get("hint", ""),
                            key=f"wc_{_wc_qk}_custom", label_visibility="collapsed",
                        )
                    else:
                        _wc_value = _wc_chosen or ""
                    _wc_ready = bool(_wc_value)

                elif _wc_type == "textarea":
                    _wc_value = st.text_area(
                        "Précisez", placeholder=_wc_q.get("hint", ""), height=100,
                        key=f"wc_{_wc_qk}_text", label_visibility="collapsed",
                    )
                    _wc_ready = True  # optionnel

                st.markdown("<div style='height:4px'></div>", unsafe_allow_html=True)
                _prev_col, _next_col = st.columns([1, 2])
                with _prev_col:
                    if _wc_idx > 0 and st.button("← Retour", key=f"wc_prev_{_wc_idx}", use_container_width=True):
                        st.session_state.ma_wc_idx = _wc_idx - 1
                        st.rerun()
                with _next_col:
                    _is_last = (_wc_idx == _wc_n - 1)
                    _btn_lbl = "🔍 Lancer la recherche de cibles" if _is_last else "Suivant →"
                    if _wc_ready and st.button(_btn_lbl, type="primary", key=f"wc_next_{_wc_idx}", use_container_width=True):
                        _new_wc_vars = {**_wc_vars, _wc_qk: _wc_value}
                        st.session_state.ma_wc_vars = _new_wc_vars
                        if not _is_last:
                            st.session_state.ma_wc_idx = _wc_idx + 1
                        else:
                            # Fusionner avec les variables positionnement
                            _all_vars = {**st.session_state.get("ma_variables", {}), **_new_wc_vars}
                            _cats_raw = _all_vars.get("categories", [])
                            if isinstance(_cats_raw, str):
                                _cat_list = [c.strip() for c in _cats_raw.split("\n") if c.strip()]
                            else:
                                _cat_list = [c for c in _cats_raw if c]
                            st.session_state["ma_variables"] = _all_vars
                            st.session_state["ma_cibles_results"] = {}
                            if _cat_list:
                                st.session_state["ma_current_category"] = _cat_list[0]
                                st.session_state["ma_categories_queue"] = _cat_list[1:]
                                st.session_state.ma_phase = "run_cibles"
                            else:
                                st.session_state.ma_phase = "done"
                            # Réinitialiser le wizard pour la prochaine fois
                            st.session_state.pop("ma_wc_idx", None)
                            st.session_state.pop("ma_wc_vars", None)
                        st.rerun()

    # ── Phase : Recherche cibles ────────────────────────────────────────────
    elif ma_phase == "run_cibles":
        variables    = st.session_state.get("ma_variables", {})
        current_cat  = st.session_state.get("ma_current_category", "")
        context_docs = st.session_state.get("ma_context_docs", "")

        _cibles_parts = []
        if variables.get("positionnement"):
            _cibles_parts.append(f"Positionnement : {variables['positionnement']}")
        _cibles_parts.append(f"Catégorie cible : {current_cat}")
        if variables.get("zone_geo"):
            _cibles_parts.append(f"Zone géographique : {variables['zone_geo']}")
        if variables.get("fourchette_ca"):
            _cibles_parts.append(f"Fourchette CA : {variables['fourchette_ca']}")
        if variables.get("exclusions"):
            _cibles_parts.append(f"Entreprises à exclure :\n{variables['exclusions']}")
        if context_docs:
            _cibles_parts.append(context_docs)
        _cibles_input = "\n".join(_cibles_parts)

        _vars_with_cat = {**variables, "categories": current_cat}
        _run_module_s4(
            "buy_03_recherche_cibles", _cibles_input,
            f"Recherche de cibles — {current_cat}",
            "check_cibles", _vars_with_cat,
        )

    # ── Phase : Satisfaction cibles ─────────────────────────────────────────
    elif ma_phase == "check_cibles":
        current_cat    = st.session_state.get("ma_current_category", "")
        cibles_results = st.session_state.get("ma_cibles_results", {})
        _raw_cibles    = st.session_state.get("ma_result_buy_03_recherche_cibles", "")

        # Sauvegarder dans le dict par catégorie
        if current_cat and current_cat not in cibles_results and _raw_cibles:
            cibles_results[current_cat] = _raw_cibles
            st.session_state["ma_cibles_results"] = cibles_results

        _cibles_text = cibles_results.get(current_cat, _raw_cibles)
        _cat_key = f"cibles_{current_cat.replace(' ', '_')[:24]}"

        _result_card(_cat_key, f"Cibles — {current_cat}", _cibles_text)

        _queue = st.session_state.get("ma_categories_queue", [])
        st.markdown('<div style="height:8px"></div>', unsafe_allow_html=True)

        if _queue:
            _next_cat = _queue[0]
            with st.container(border=True):
                st.markdown(
                    f'<div style="font-size:0.88rem;font-weight:600;margin-bottom:6px;">Continuer sur la catégorie suivante ?</div>'
                    f'<div style="font-size:0.83rem;color:#444444;margin-bottom:10px;">→ {_next_cat}</div>',
                    unsafe_allow_html=True,
                )
                _c1, _c2, _c3 = st.columns(3)
                with _c1:
                    if st.button(f"→ Lancer {_next_cat[:20]}", type="primary", use_container_width=True, key="cibles_next_yes"):
                        st.session_state["ma_current_category"] = _next_cat
                        st.session_state["ma_categories_queue"] = _queue[1:]
                        st.session_state.pop("ma_result_buy_03_recherche_cibles", None)
                        st.session_state.ma_phase = "run_cibles"
                        st.rerun()
                with _c2:
                    if st.button("↺ Relancer cette catégorie", use_container_width=True, key="cibles_rerun"):
                        st.session_state.pop("ma_result_buy_03_recherche_cibles", None)
                        cibles_results.pop(current_cat, None)
                        st.session_state["ma_cibles_results"] = cibles_results
                        st.session_state.ma_phase = "run_cibles"
                        st.rerun()
                with _c3:
                    if st.button("✓ Terminer la mission", use_container_width=True, key="cibles_done"):
                        st.session_state.ma_phase = "done"
                        st.rerun()
        else:
            _c1, _c2 = st.columns(2)
            with _c1:
                if st.button("↺ Relancer cette catégorie", use_container_width=True, key="cibles_rerun_last"):
                    st.session_state.pop("ma_result_buy_03_recherche_cibles", None)
                    cibles_results.pop(current_cat, None)
                    st.session_state["ma_cibles_results"] = cibles_results
                    st.session_state.ma_phase = "run_cibles"
                    st.rerun()
            with _c2:
                if st.button("✓ Terminer la mission", type="primary", use_container_width=True, key="cibles_done_last"):
                    st.session_state.ma_phase = "done"
                    st.rerun()

    # ── Phase : Done — résumé complet ──────────────────────────────────────
    elif ma_phase == "done":
        cibles_results = st.session_state.get("ma_cibles_results", {})
        result_v = st.session_state.get("ma_result_buy_01_carto_verticale", "")
        result_h = st.session_state.get("ma_result_buy_02_carto_horizontale", "")

        st.markdown('<div style="font-size:0.88rem;font-weight:700;color:#065F46;margin-bottom:12px;">✓ Mission terminée</div>', unsafe_allow_html=True)

        if result_v:
            _result_card("buy_01_carto_verticale", "Cartographie verticale", result_v)
            st.markdown('<div style="height:6px"></div>', unsafe_allow_html=True)
        if result_h:
            _result_card("buy_02_carto_horizontale", "Cartographie horizontale", result_h)
            st.markdown('<div style="height:6px"></div>', unsafe_allow_html=True)
        for _cat, _txt in cibles_results.items():
            _ck = f"cibles_{_cat.replace(' ', '_')[:24]}_done"
            _result_card(_ck, f"Cibles — {_cat}", _txt)
            st.markdown('<div style="height:6px"></div>', unsafe_allow_html=True)

        # Export global
        if result_v or result_h or cibles_results:
            st.markdown('<div style="height:4px"></div>', unsafe_allow_html=True)
            with st.container(border=True):
                st.markdown(
                    '<div style="font-size:0.95rem;font-weight:700;color:#111111;margin-bottom:4px;">📥 Rapport complet</div>'
                    '<div style="font-size:0.82rem;color:#6B7280;margin-bottom:12px;">Tous les résultats dans un seul fichier Excel</div>',
                    unsafe_allow_html=True,
                )
                try:
                    from export_ma_xlsx import generate_ma_xlsx
                    _all_cibles = "\n\n".join(
                        f"## {_c}\n{_t}" for _c, _t in cibles_results.items()
                    ) if cibles_results else ""
                    _xlsx_global = generate_ma_xlsx(
                        company=ma_company,
                        text_v=result_v,
                        text_h=result_h,
                        text_cibles=_all_cibles,
                    )
                    _fname_g = f"screening_{ma_company.replace(' ', '_').lower()}.xlsx"
                    st.download_button(
                        "📥 Télécharger en Excel",
                        data=_xlsx_global,
                        file_name=_fname_g,
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True,
                        type="primary",
                    )
                except Exception as _xe:
                    st.warning(f"Export Excel indisponible : {_xe}")

        st.markdown('<div style="height:8px"></div>', unsafe_allow_html=True)
        if st.button("🔄 Recommencer la mission", use_container_width=True):
            _s4_quit()
            st.rerun()
