"""
Génère un Excel de screening M&A calqué sur le template _Template Screening.xlsm.
Feuilles produites : Mapping V, Mapping H, Cibles
"""
import io
import re
import anthropic
import os
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter

# ─── Couleurs template ───────────────────────────────────────────────────────
C_HEADER_FILL  = "1F3864"   # bleu marine foncé (couleur template)
C_HEADER_FONT  = "FFFFFF"   # blanc
C_TITLE_FILL   = "2E4A8B"   # bleu titre
C_ROW_ODD      = "EEF2F8"   # bleu très clair rangée impaire
C_ROW_EVEN     = "FFFFFF"   # blanc rangée paire
C_BORDER       = "B8C4D4"   # gris-bleu border
C_LABEL_FILL   = "F0F4FA"   # fond libellé col A

THIN = Side(style="thin", color=C_BORDER)
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def _hdr_cell(ws, row, col, value, width=None):
    c = ws.cell(row=row, column=col, value=value)
    c.font      = Font(name="Arial", bold=True, color=C_HEADER_FONT, size=9)
    c.fill      = PatternFill("solid", fgColor=C_HEADER_FILL)
    c.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
    c.border    = BORDER
    if width:
        ws.column_dimensions[get_column_letter(col)].width = width
    return c


def _data_cell(ws, row, col, value, bold=False, odd=True, wrap=True):
    c = ws.cell(row=row, column=col, value=value)
    c.font      = Font(name="Arial", bold=bold, size=9)
    c.fill      = PatternFill("solid", fgColor=C_ROW_ODD if odd else C_ROW_EVEN)
    c.alignment = Alignment(wrap_text=wrap, vertical="top")
    c.border    = BORDER
    return c


def _title_row(ws, row, text, ncols):
    c = ws.cell(row=row, column=2, value=text)
    c.font      = Font(name="Arial", bold=True, color=C_HEADER_FONT, size=12)
    c.fill      = PatternFill("solid", fgColor=C_TITLE_FILL)
    c.alignment = Alignment(horizontal="left", vertical="center")
    if ncols > 1:
        ws.merge_cells(
            start_row=row, start_column=2, end_row=row, end_column=ncols + 1
        )
    ws.row_dimensions[row].height = 28


# ─── Parseurs ────────────────────────────────────────────────────────────────

def _parse_with_ai(text: str, schema_desc: str, example_json: str) -> list[dict]:
    """Utilise Claude Haiku pour extraire du JSON structuré depuis du markdown."""
    client = anthropic.Anthropic(api_key=os.environ["ANTHROPIC_API_KEY"])
    prompt = f"""Extrait les données du texte suivant et retourne UNIQUEMENT un JSON valide (liste d'objets).

Schéma attendu pour chaque objet :
{schema_desc}

Exemple de format :
{example_json}

Texte à analyser :
{text[:12000]}

Réponds UNIQUEMENT avec le JSON (tableau), sans texte autour."""
    resp = client.messages.create(
        model="claude-haiku-4-5-20251001",
        max_tokens=4000,
        messages=[{"role": "user", "content": prompt}],
    )
    raw = resp.content[0].text.strip()
    m = re.search(r"\[.*\]", raw, re.DOTALL)
    if m:
        import json
        try:
            return json.loads(m.group())
        except Exception:
            pass
    return []


def _parse_mapping_v(text: str, company: str) -> list[dict]:
    schema = """{
  "num": "V1",
  "categorie": "Libellé de la catégorie",
  "definition": "Définition & positionnement",
  "lien_acquéreur": "Lien avec l'acquéreur",
  "dans_categorie": "L'acquéreur dans cette catégorie ? Oui/Non + explication",
  "exemples": "Exemples d'acteurs (1-2)",
  "pourquoi_important": "Pourquoi c'est important",
  "appetit_acquisition": "Pourquoi ils pourraient acquérir / note d'appétit"
}"""
    example = '[{"num":"V1","categorie":"Fabricants...","definition":"...","lien_acquéreur":"...","dans_categorie":"Non...","exemples":"Acteur A, Acteur B","pourquoi_important":"...","appetit_acquisition":"2/3"}]'
    return _parse_with_ai(text, schema, example)


def _parse_mapping_h(text: str, company: str) -> list[dict]:
    schema = """{
  "num": "H1",
  "segment": "Libellé du segment concurrentiel",
  "definition": "Définition & frontières du segment",
  "rivalite": "Rivalité vs l'acquéreur",
  "differentiation": "Différenciation clé dans le segment",
  "dans_segment": "L'acquéreur dans ce segment ? Oui/Non",
  "exemples": "Exemples d'acteurs (1-2)",
  "pourquoi_important": "Pourquoi c'est important",
  "sources": "Sources"
}"""
    example = '[{"num":"H1","segment":"Réseaux de...","definition":"...","rivalite":"Directe","differentiation":"...","dans_segment":"Oui","exemples":"Acteur A","pourquoi_important":"...","sources":"Site corporate"}]'
    return _parse_with_ai(text, schema, example)


def _parse_cibles(text: str, company: str) -> list[dict]:
    schema = """{
  "liste": "Liste 1 / Liste 2 / Liste 3",
  "societe": "Nom de la société ou groupe",
  "pays": "Pays",
  "positionnement": "Positionnement (catégorie V ou H)",
  "metier": "Métier",
  "offre": "Offre",
  "clients": "Clients types",
  "modele": "Modèle commercial",
  "chiffres_cles": "Chiffres clés (CA, EBITDA, effectifs)",
  "actionnariat": "Actionnariat & gouvernance",
  "operations_ma": "Opérations M&A récentes",
  "sources": "Sources utilisées",
  "positionnement_vs": "Positionnement par rapport à l'acquéreur"
}"""
    example = '[{"liste":"Liste 1","societe":"Acme SAS","pays":"France","positionnement":"V4 — Intégrateurs","metier":"Intégrateur","offre":"...","clients":"B2B industriels","modele":"Vente directe","chiffres_cles":"CA 8 M€, 45 pers.","actionnariat":"Familial","operations_ma":"Aucune","sources":"Site corporate","positionnement_vs":"Concurrent direct sur segment automatisme"}]'
    return _parse_with_ai(text, schema, example)


# ─── Constructeurs de feuilles ────────────────────────────────────────────────

def _build_mapping_v(wb: Workbook, rows: list[dict], company: str):
    ws = wb.create_sheet("Mapping V")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 3

    _title_row(ws, 2, "Mapping vertical", 9)
    ws.row_dimensions[2].height = 28

    headers = [
        ("#", 5), ("Catégorie d'acteurs", 28), ("Définition & positionnement", 52),
        (f"Lien avec {company}", 45), (f"{company} dans cette catégorie ?", 35),
        ("Exemples d'acteurs (1–2)", 38), ("Pourquoi c'est important", 48),
        ("Appétit acquisition / note", 48), ("Note", 10),
    ]
    for ci, (label, w) in enumerate(headers, start=2):
        _hdr_cell(ws, 6, ci, label, w)
    ws.row_dimensions[6].height = 36

    for ri, row in enumerate(rows, start=7):
        odd = (ri % 2 == 1)
        _data_cell(ws, ri, 2, row.get("num", ""), bold=True, odd=odd, wrap=False)
        _data_cell(ws, ri, 3, row.get("categorie", ""), bold=True, odd=odd)
        _data_cell(ws, ri, 4, row.get("definition", ""), odd=odd)
        _data_cell(ws, ri, 5, row.get("lien_acquéreur", ""), odd=odd)
        _data_cell(ws, ri, 6, row.get("dans_categorie", ""), odd=odd)
        _data_cell(ws, ri, 7, row.get("exemples", ""), odd=odd)
        _data_cell(ws, ri, 8, row.get("pourquoi_important", ""), odd=odd)
        _data_cell(ws, ri, 9, row.get("appetit_acquisition", ""), odd=odd)
        ws.row_dimensions[ri].height = 90


def _build_mapping_h(wb: Workbook, rows: list[dict], company: str):
    ws = wb.create_sheet("Mapping H")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 3

    _title_row(ws, 2, "Mapping horizontal", 9)

    headers = [
        ("#", 5), ("Segment concurrentiel", 30), ("Définition & frontières", 50),
        (f"Rivalité vs {company}", 40), ("Différenciation clé", 40),
        (f"{company} dans ce segment ?", 35), ("Exemples d'acteurs (1–2)", 38),
        ("Pourquoi c'est important", 50), ("Sources", 30),
    ]
    for ci, (label, w) in enumerate(headers, start=2):
        _hdr_cell(ws, 6, ci, label, w)
    ws.row_dimensions[6].height = 36

    for ri, row in enumerate(rows, start=7):
        odd = (ri % 2 == 1)
        _data_cell(ws, ri, 2, row.get("num", ""), bold=True, odd=odd, wrap=False)
        _data_cell(ws, ri, 3, row.get("segment", ""), bold=True, odd=odd)
        _data_cell(ws, ri, 4, row.get("definition", ""), odd=odd)
        _data_cell(ws, ri, 5, row.get("rivalite", ""), odd=odd)
        _data_cell(ws, ri, 6, row.get("differentiation", ""), odd=odd)
        _data_cell(ws, ri, 7, row.get("dans_segment", ""), odd=odd)
        _data_cell(ws, ri, 8, row.get("exemples", ""), odd=odd)
        _data_cell(ws, ri, 9, row.get("pourquoi_important", ""), odd=odd)
        _data_cell(ws, ri, 10, row.get("sources", ""), odd=odd)
        ws.row_dimensions[ri].height = 90


def _build_cibles(wb: Workbook, rows: list[dict], company: str):
    ws = wb.create_sheet("Cibles")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 3

    _title_row(ws, 2, "Long-list de cibles", 14)

    headers = [
        ("Liste", 10), ("Société / Groupe", 24), ("Pays", 12),
        ("Positionnement", 30), ("Métier", 30), ("Offre", 52),
        ("Clients", 45), ("Modèle", 40), ("Chiffres clés", 36),
        ("Actionnariat & gouvernance", 42), ("Opérations M&A", 44),
        ("Sources", 30), (f"Positionnement vs {company}", 55),
    ]
    for ci, (label, w) in enumerate(headers, start=2):
        _hdr_cell(ws, 7, ci, label, w)
    ws.row_dimensions[7].height = 36

    for ri, row in enumerate(rows, start=8):
        odd = (ri % 2 == 0)
        _data_cell(ws, ri, 2, row.get("liste", ""), odd=odd, wrap=False)
        _data_cell(ws, ri, 3, row.get("societe", ""), bold=True, odd=odd)
        _data_cell(ws, ri, 4, row.get("pays", ""), odd=odd, wrap=False)
        _data_cell(ws, ri, 5, row.get("positionnement", ""), odd=odd)
        _data_cell(ws, ri, 6, row.get("metier", ""), bold=True, odd=odd)
        _data_cell(ws, ri, 7, row.get("offre", ""), odd=odd)
        _data_cell(ws, ri, 8, row.get("clients", ""), odd=odd)
        _data_cell(ws, ri, 9, row.get("modele", ""), odd=odd)
        _data_cell(ws, ri, 10, row.get("chiffres_cles", ""), odd=odd)
        _data_cell(ws, ri, 11, row.get("actionnariat", ""), odd=odd)
        _data_cell(ws, ri, 12, row.get("operations_ma", ""), odd=odd)
        _data_cell(ws, ri, 13, row.get("sources", ""), odd=odd)
        _data_cell(ws, ri, 14, row.get("positionnement_vs", ""), odd=odd)
        ws.row_dimensions[ri].height = 110


# ─── Point d'entrée public ────────────────────────────────────────────────────

def generate_ma_xlsx(
    company: str,
    text_v: str = "",
    text_h: str = "",
    text_cibles: str = "",
) -> bytes:
    """
    Génère le classeur Excel complet (Mapping V + H + Cibles).
    Retourne les bytes du fichier .xlsx.
    """
    wb = Workbook()
    wb.remove(wb.active)   # supprimer la feuille vide par défaut

    if text_v:
        rows_v = _parse_mapping_v(text_v, company)
        if rows_v:
            _build_mapping_v(wb, rows_v, company)

    if text_h:
        rows_h = _parse_mapping_h(text_h, company)
        if rows_h:
            _build_mapping_h(wb, rows_h, company)

    if text_cibles:
        rows_c = _parse_cibles(text_cibles, company)
        if rows_c:
            _build_cibles(wb, rows_c, company)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
